from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_wtf.csrf import CSRFProtect, CSRFError
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
from functools import wraps
import secrets
import datetime
from flask_mail import Mail, Message
try:
    from authlib.integrations.flask_client import OAuth
except ImportError:
    OAuth = None
import json
import requests
from sqlalchemy import inspect
from sqlalchemy import or_
from dotenv import load_dotenv

# Base directory of the project (parent of backend/)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))

# Load environment variables from .env file (project root)
load_dotenv(os.path.join(BASE_DIR, '.env'))

app = Flask(
    __name__,
    template_folder=os.path.join(BASE_DIR, 'frontend', 'templates'),
    static_folder=os.path.join(BASE_DIR, 'frontend', 'static')
)
# Read from environment in production; allow a dev-only fallback for local runs.
_secret_key = os.environ.get('SECRET_KEY')
_is_dev = app.debug or os.environ.get('FLASK_ENV') == 'development'
if _secret_key:
    app.config['SECRET_KEY'] = _secret_key
elif _is_dev:
    app.config['SECRET_KEY'] = 'dev-secret-key-change-this-in-prod-982374923'
else:
    raise RuntimeError("SECRET_KEY must be set in production")

# Initialize CSRF protection
app.config['WTF_CSRF_TIME_LIMIT'] = 86400  # 24 hours
csrf = CSRFProtect(app)

# Render persistent disk path (must be defined before upload/DB config)
RENDER_DATA_DIR = os.environ.get('RENDER_DATA_DIR', '/opt/render/project/data')

# File upload configuration
if os.environ.get('RENDER'):
    UPLOAD_FOLDER = os.path.join(RENDER_DATA_DIR, 'uploads')
elif os.environ.get('VERCEL'):
    UPLOAD_FOLDER = '/tmp/uploads'
else:
    UPLOAD_FOLDER = os.path.join(BACKEND_DIR, 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'ppt', 'pptx', 'xls', 'xlsx', 'txt', 'zip', 'rar', 'png', 'jpg', 'jpeg'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB max upload size
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
# Database configuration
# Priority: DATABASE_URL env var (Supabase/PostgreSQL) > SQLite fallback
_database_url = os.environ.get('DATABASE_URL')
if _database_url:
    # Supabase / external PostgreSQL
    # Fix Heroku/Supabase URI scheme: postgres:// → postgresql://
    if _database_url.startswith('postgres://'):
        _database_url = _database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = _database_url
elif os.environ.get('RENDER'):
    db_path = os.path.join(RENDER_DATA_DIR, 'fyp.db')
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///fyp.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Database connection settings
_engine_opts = {'pool_pre_ping': True}
if 'sqlite' in app.config.get('SQLALCHEMY_DATABASE_URI', ''):
    _engine_opts['connect_args'] = {'check_same_thread': False}
else:
    # PostgreSQL connection pool settings
    _engine_opts['pool_size'] = 5
    _engine_opts['max_overflow'] = 10
    _engine_opts['pool_recycle'] = 300
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = _engine_opts

# Security configuration for session cookies
# Secure cookies in production (HTTPS), allow HTTP for local dev.
app.config['SESSION_COOKIE_SECURE'] = not _is_dev
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Allow OAuth over HTTP for local testing
if app.debug or os.environ.get('FLASK_ENV') == 'development':
    os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
    print("[INFO] OAuth insecure transport enabled for local development")

# Email configuration
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', '587'))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'true').lower() == 'true'
app.config['MAIL_USE_SSL'] = os.environ.get('MAIL_USE_SSL', 'false').lower() == 'true'
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get(
    'MAIL_DEFAULT_SENDER',
    app.config['MAIL_USERNAME'] or 'no-reply@example.com'
)
app.config['MAIL_MAX_EMAILS'] = None
app.config['MAIL_ASCII_ATTACHMENTS'] = False
app.config['MAIL_SUPPRESS_SEND'] = not (
    app.config['MAIL_USERNAME'] and app.config['MAIL_PASSWORD']
)
app.config['MAIL_DEBUG'] = app.debug

# OAuth Configuration
# Load from environment variables (.env file or system environment)
app.config['GOOGLE_CLIENT_ID'] = os.environ.get('GOOGLE_CLIENT_ID', 'not-configured')
app.config['GOOGLE_CLIENT_SECRET'] = os.environ.get('GOOGLE_CLIENT_SECRET', 'not-configured')
app.config['GOOGLE_DISCOVERY_URL'] = 'https://accounts.google.com/.well-known/openid-configuration'

# Check if OAuth is properly configured
OAUTH_CONFIGURED = (app.config['GOOGLE_CLIENT_ID'] != 'not-configured' and 
                    app.config['GOOGLE_CLIENT_SECRET'] != 'not-configured')

# Initialize extensions
db = SQLAlchemy(app)
mail = Mail(app)
if OAuth:
    oauth = OAuth(app)
else:
    oauth = None

# Create a directory for storing email files
EMAIL_DIR = os.path.join(BACKEND_DIR, 'emails')
os.makedirs(EMAIL_DIR, exist_ok=True)

# Helper function to send emails reliably
def send_email(to_email, subject, body, from_email=None):
    """
    Sends an email using multiple methods, falling back as needed.
    For development, it writes the email to a file.
    """
    if from_email is None:
        from_email = app.config['MAIL_DEFAULT_SENDER']
    
    email_content = {
        'to': to_email,
        'from': from_email,
        'subject': subject,
        'body': body,
        'timestamp': str(datetime.datetime.now())
    }
    
    # Always save to file (for logging purposes)
    filename = f"email_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_{to_email.replace('@', '_').replace('.', '_')}.json"
    file_path = os.path.join(EMAIL_DIR, filename)
    
    with open(file_path, 'w') as f:
        json.dump(email_content, f, indent=4)
    
    print(f"Email saved to {file_path}")
    
    # Try to send using the Brevo/Sendinblue API (free tier)
    try:
        # Use the Brevo/Sendinblue API (formerly Sendinblue)
        api_url = "https://api.brevo.com/v3/smtp/email"
        
        # Note: You would need to sign up for a free Brevo account
        # and replace this API key with your own
        headers = {
            "accept": "application/json",
            "content-type": "application/json",
            "api-key": "YOUR_API_KEY_HERE"  # Replace with your actual API key
        }
        
        payload = {
            "sender": {"email": from_email, "name": "FYP Management System"},
            "to": [{"email": to_email}],
            "subject": subject,
            "htmlContent": body.replace('\n', '<br>')
        }
        
        # Only attempt to send via API if the API key is properly set
        if headers["api-key"] != "YOUR_API_KEY_HERE":
            response = requests.post(api_url, json=payload, headers=headers)
            if response.status_code == 201:
                print(f"Email sent successfully via Brevo API to {to_email}")
                return True, "Email sent via Brevo API"
            else:
                print(f"Brevo API error: {response.status_code} - {response.text}")
                # Fall back to Flask-Mail
        
    except Exception as e:
        print(f"Error with Brevo API: {str(e)}")
    
    # If we're in debug mode, don't attempt SMTP which will likely timeout
    if app.debug:
        return True, f"Email saved to file: {file_path}"
    
    # Try Flask-Mail as a last resort (production only)
    try:
        msg = Message(subject, recipients=[to_email], body=body, sender=from_email)
        mail.send(msg)
        return True, "Email sent via Flask-Mail"
    except Exception as e:
        error_msg = f"Flask-Mail error: {str(e)}"
        print(error_msg)
        return False, error_msg

# Test mail connection on startup if in debug mode
if app.debug and not os.environ.get('VERCEL'):
    with app.app_context():
        try:
            mail.connect()
            print("Mail server connection successful!")
        except Exception as e:
            print(f"Warning: Could not connect to mail server: {str(e)}")
            print("Password reset emails will not be sent, but the app will still function.")
            
# Setup Google OAuth (only if properly configured)
if OAUTH_CONFIGURED and oauth:
    try:
        google = oauth.register(
            name='google',
            client_id=app.config['GOOGLE_CLIENT_ID'],
            client_secret=app.config['GOOGLE_CLIENT_SECRET'],
            server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
            client_kwargs={'scope': 'openid email profile'}
        )
        print("[OK] Google OAuth properly configured")
    except Exception as e:
        print(f"[ERROR] Error initializing Google OAuth: {str(e)}")
        google = None
else:
    google = None
    if not OAUTH_CONFIGURED:
        print("[WARNING] Google OAuth not configured. Set GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET environment variables.")
        print("  See GOOGLE_OAUTH_SETUP.md for instructions.")

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256))
    first_name = db.Column(db.String(50))
    last_name = db.Column(db.String(50))
    role = db.Column(db.String(20))
    google_id = db.Column(db.String(100), unique=True, nullable=True)
    reset_token = db.Column(db.String(100), nullable=True)
    reset_token_expiry = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())

    # Extended profile fields
    program = db.Column(db.String(50))      # For Students (CS, IT, AI)
    semester = db.Column(db.String(20))     # For Students
    
    highest_degree = db.Column(db.String(50)) # For Supervisors
    specialization = db.Column(db.String(100)) # For Supervisors
    affiliation = db.Column(db.String(100))    # For Supervisors (NUTECH vs Other)
    other_affiliation = db.Column(db.String(100)) # For Supervisors

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        if not self.password_hash:
            return False
        return check_password_hash(self.password_hash, password)
        
    def generate_reset_token(self):
        self.reset_token = secrets.token_hex(16)
        self.reset_token_expiry = datetime.datetime.now() + datetime.timedelta(hours=1)
        db.session.commit()
        return self.reset_token
        
    def verify_reset_token(self, token):
        if self.reset_token != token:
            return False
        if datetime.datetime.now() > self.reset_token_expiry:
            return False
        return True
        
    def clear_reset_token(self):
        self.reset_token = None
        self.reset_token_expiry = None
        db.session.commit()

class LoginAttempt(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), nullable=False)
    success = db.Column(db.Boolean, default=False)
    ip_address = db.Column(db.String(50))
    user_agent = db.Column(db.String(255))
    timestamp = db.Column(db.DateTime, default=db.func.current_timestamp())
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    
    user = db.relationship('User', backref='login_attempts')
    
    def __repr__(self):
        return f"<LoginAttempt {self.email} at {self.timestamp}>"

class StudentGroup(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.String(10), unique=True, nullable=False)
    project_title = db.Column(db.String(200), nullable=False)
    project_description = db.Column(db.Text, nullable=True)
    supervisor_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    updated_at = db.Column(db.DateTime, default=db.func.current_timestamp(), onupdate=db.func.current_timestamp())
    supervisor = db.relationship('User', backref='supervised_groups')
    remarks = db.relationship('Remark', backref='group', lazy='dynamic')
    
class ProjectStatus(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    group_id = db.Column(db.Integer, db.ForeignKey('student_group.id'), nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    status = db.Column(db.String(50), default='Pending')  # 'Pending', 'Accepted', 'Conditionally Accepted', 'Deferred'
    feedback = db.Column(db.Text, nullable=True)
    student_feedback = db.Column(db.Text, nullable=True)  # Brief feedback for students
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    updated_at = db.Column(db.DateTime, default=db.func.current_timestamp(), onupdate=db.func.current_timestamp())
    group = db.relationship('StudentGroup', backref='statuses')
    teacher = db.relationship('User', backref='project_statuses')

class ProjectProposal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    major = db.Column(db.String(50), nullable=False)  # AI/ML, Cyber Security, Blockchain, Web Development, etc.
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    status = db.Column(db.String(20), default='Pending')  # 'Pending', 'Approved', 'Rejected'
    feedback = db.Column(db.Text)
    
    # Foreign Keys
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    supervisor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # Optional at first
    admin_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    
    # Relationships
    student = db.relationship('User', foreign_keys=[student_id], backref='proposed_projects')
    supervisor = db.relationship('User', foreign_keys=[supervisor_id], backref='received_proposals')
    admin = db.relationship('User', foreign_keys=[admin_id], backref='reviewed_proposals')
    
    def __repr__(self):
        return f"Proposal: {self.title} by {self.student.first_name} {self.student.last_name}"

class ProjectMilestone(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    due_date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), default='Pending')  # 'Pending', 'Completed', 'Late'
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    
    # Foreign Keys
    group_id = db.Column(db.Integer, db.ForeignKey('student_group.id'), nullable=False)
    
    # Relationships
    group = db.relationship('StudentGroup', backref='milestones')
    
    def __repr__(self):
        return f"Milestone: {self.title} for Group: {self.group.group_id}"

class ProjectDetails(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    description = db.Column(db.Text)
    major = db.Column(db.String(50))  # AI/ML, Cyber Security, Blockchain, Web Development, etc.
    progress = db.Column(db.Integer, default=0)  # Percentage of completion
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)
    
    # Foreign Keys
    group_id = db.Column(db.Integer, db.ForeignKey('student_group.id'), nullable=False, unique=True)
    
    # Relationships
    group = db.relationship('StudentGroup', backref=db.backref('details', uselist=False))
    
    def __repr__(self):
        return f"Details for Group: {self.group.group_id}"

class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    message = db.Column(db.Text, nullable=False)
    notification_type = db.Column(db.String(50))
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    
    user = db.relationship('User', backref='notifications')
    
class Remark(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    content = db.Column(db.Text, nullable=False)
    timestamp = db.Column(db.DateTime, nullable=False, default=db.func.current_timestamp())
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    group_id = db.Column(db.Integer, db.ForeignKey('student_group.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # NULL = whole group, set = specific student
    teacher = db.relationship('User', foreign_keys=[teacher_id], backref='remarks')
    student = db.relationship('User', foreign_keys=[student_id], backref='targeted_remarks')

class TeacherUsername(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    is_used = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    user = db.relationship('User', backref='teacher_username')

class GroupMember(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    group_id = db.Column(db.Integer, db.ForeignKey('student_group.id'), nullable=False)
    user = db.relationship('User', backref='group_memberships')
    group = db.relationship('StudentGroup', backref='members')

# Scheduling Models
class TimeSlot(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    day = db.Column(db.String(10), nullable=False)  # Monday, Tuesday, etc.
    start_time = db.Column(db.String(10), nullable=False)  # HH:MM format
    end_time = db.Column(db.String(10), nullable=False)  # HH:MM format
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    
    # Relationships
    teacher_schedules = db.relationship('TeacherSchedule', backref='time_slot', lazy='dynamic')
    room_schedules = db.relationship('RoomSchedule', backref='time_slot', lazy='dynamic')
    
    def __repr__(self):
        return f"{self.day} {self.start_time}-{self.end_time}"

class Room(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(20), nullable=False, unique=True)  # HL1, HL2, etc.
    capacity = db.Column(db.Integer, default=30)
    description = db.Column(db.String(200))
    
    # Relationships
    schedules = db.relationship('RoomSchedule', backref='room', lazy='dynamic')
    
    def __repr__(self):
        return self.name

class TeacherSchedule(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    time_slot_id = db.Column(db.Integer, db.ForeignKey('time_slot.id'), nullable=False)
    group_id = db.Column(db.Integer, db.ForeignKey('student_group.id'), nullable=True)
    subject = db.Column(db.String(100))
    class_name = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    
    # The teacher should be unique for each time slot (a teacher can't teach two classes at the same time)
    __table_args__ = (db.UniqueConstraint('teacher_id', 'time_slot_id', name='uix_teacher_timeslot'),)
    
    # Relationships
    teacher = db.relationship('User', backref='teaching_schedules')
    group = db.relationship('StudentGroup', backref='schedules')

class RoomSchedule(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    room_id = db.Column(db.Integer, db.ForeignKey('room.id'), nullable=False)
    time_slot_id = db.Column(db.Integer, db.ForeignKey('time_slot.id'), nullable=False)
    group_id = db.Column(db.Integer, db.ForeignKey('student_group.id'), nullable=True)
    class_name = db.Column(db.String(100))  # Name of the class/course
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    
    # Each room should have unique time slots (a room can't have two classes at the same time)
    __table_args__ = (db.UniqueConstraint('room_id', 'time_slot_id', name='uix_room_timeslot'),)
    
    # Relationships
    group = db.relationship('StudentGroup', backref='room_schedule_entries')

class Viva(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    scheduled_date = db.Column(db.Date, nullable=False)
    scheduled_time = db.Column(db.String(10), nullable=False)  # HH:MM format
    duration_minutes = db.Column(db.Integer, default=30)
    location = db.Column(db.String(100))
    status = db.Column(db.String(20), default='Scheduled')  # 'Scheduled', 'Completed', 'Cancelled'
    notes = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=db.func.current_timestamp())
    
    # Foreign Keys
    group_id = db.Column(db.Integer, db.ForeignKey('student_group.id'), nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # Relationships
    group = db.relationship('StudentGroup', backref='vivas')
    teacher = db.relationship('User', backref='assigned_vivas')
    
    def __repr__(self):
        return f"Viva for Group {self.group.group_id} on {self.scheduled_date} at {self.scheduled_time}"

class Submission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=True)
    filename = db.Column(db.String(300), nullable=False)  # stored filename on disk
    original_filename = db.Column(db.String(300), nullable=False)  # original user filename
    file_size = db.Column(db.Integer, default=0)  # size in bytes
    file_type = db.Column(db.String(20))  # extension
    submission_type = db.Column(db.String(50), default='General')  # 'Proposal', 'Progress Report', 'Final Report', 'Code', 'Presentation', 'General'
    status = db.Column(db.String(20), default='Submitted')  # 'Submitted', 'Reviewed', 'Approved', 'Rejected'
    feedback = db.Column(db.Text, nullable=True)  # supervisor/faculty feedback
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

    # Foreign Keys
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    group_id = db.Column(db.Integer, db.ForeignKey('student_group.id'), nullable=False)

    # Relationships
    student = db.relationship('User', backref='submissions')
    group = db.relationship('StudentGroup', backref='submissions')

    def __repr__(self):
        return f"Submission: {self.title} by Student {self.student_id}"

    @property
    def file_size_display(self):
        if self.file_size < 1024:
            return f"{self.file_size} B"
        elif self.file_size < 1024 * 1024:
            return f"{self.file_size / 1024:.1f} KB"
        else:
            return f"{self.file_size / (1024 * 1024):.1f} MB"

class AssignedWork(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=True)
    due_date = db.Column(db.Date, nullable=True)
    priority = db.Column(db.String(20), default='Medium')  # Low, Medium, High, Urgent
    work_type = db.Column(db.String(50), default='General')  # Task, Report, Presentation, Review, Code, Milestone, General
    status = db.Column(db.String(20), default='Pending')  # Pending, In Progress, Submitted, Completed, Overdue, Needs Revision
    student_response = db.Column(db.Text, nullable=True)  # student notes when marking done
    supervisor_comment = db.Column(db.Text, nullable=True)  # supervisor feedback/comment
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.datetime.utcnow, onupdate=datetime.datetime.utcnow)

    # Foreign Keys
    group_id = db.Column(db.Integer, db.ForeignKey('student_group.id'), nullable=False)
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # NULL = whole group
    assigned_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    submission_id = db.Column(db.Integer, db.ForeignKey('submission.id'), nullable=True)  # linked submission

    # Relationships
    group = db.relationship('StudentGroup', backref='assigned_works')
    student = db.relationship('User', foreign_keys=[assigned_to], backref='received_works')
    supervisor = db.relationship('User', foreign_keys=[assigned_by], backref='created_works')
    linked_submission = db.relationship('Submission', backref='assigned_work_ref')

    def __repr__(self):
        return f"AssignedWork: {self.title} for Group {self.group_id}"

    @property
    def is_overdue(self):
        if self.due_date and self.status not in ('Completed', 'Submitted'):
            return datetime.date.today() > self.due_date
        return False

# ============================================
# DATA INTEGRITY VALIDATORS
# ============================================

from sqlalchemy import event
import logging

logger = logging.getLogger(__name__)

# Validate Remarks before insert
@event.listens_for(Remark, 'before_insert')
@event.listens_for(Remark, 'before_update')
def validate_remark_data(mapper, connection, target):
    """Validate remark data before saving to database"""
    if not target.content or len(str(target.content).strip()) == 0:
        raise ValueError("Remark content cannot be empty")
    if not target.teacher_id:
        raise ValueError("Remark must be assigned to a teacher")
    if not target.group_id:
        raise ValueError("Remark must be assigned to a group")
    # Verify teacher exists
    teacher = User.query.get(target.teacher_id)
    if not teacher:
        raise ValueError(f"Teacher with ID {target.teacher_id} does not exist")
    # Verify group exists
    group = StudentGroup.query.get(target.group_id)
    if not group:
        raise ValueError(f"StudentGroup with ID {target.group_id} does not exist")

# Validate TeacherSchedule before insert
@event.listens_for(TeacherSchedule, 'before_insert')
@event.listens_for(TeacherSchedule, 'before_update')
def validate_teacher_schedule(mapper, connection, target):
    """Validate teacher schedule data before saving"""
    if not target.teacher_id:
        raise ValueError("Schedule must be assigned to a teacher")
    if not target.time_slot_id:
        raise ValueError("Schedule must have a time slot")
    # Verify teacher exists
    teacher = User.query.get(target.teacher_id)
    if not teacher:
        raise ValueError(f"Teacher with ID {target.teacher_id} does not exist")
    # Verify time slot exists
    slot = TimeSlot.query.get(target.time_slot_id)
    if not slot:
        raise ValueError(f"TimeSlot with ID {target.time_slot_id} does not exist")

# Validate RoomSchedule before insert
@event.listens_for(RoomSchedule, 'before_insert')
@event.listens_for(RoomSchedule, 'before_update')
def validate_room_schedule(mapper, connection, target):
    """Validate room schedule data before saving"""
    if not target.room_id:
        raise ValueError("Schedule must be assigned to a room")
    if not target.time_slot_id:
        raise ValueError("Schedule must have a time slot")
    # Verify room exists
    room = Room.query.get(target.room_id)
    if not room:
        raise ValueError(f"Room with ID {target.room_id} does not exist")
    # Verify time slot exists
    slot = TimeSlot.query.get(target.time_slot_id)
    if not slot:
        raise ValueError(f"TimeSlot with ID {target.time_slot_id} does not exist")

# Validate Viva before insert
@event.listens_for(Viva, 'before_insert')
@event.listens_for(Viva, 'before_update')
def validate_viva(mapper, connection, target):
    """Validate viva data before saving"""
    if not target.group_id:
        raise ValueError("Viva must be assigned to a group")
    if not target.teacher_id:
        raise ValueError("Viva must be assigned to a teacher")
    # Verify group exists
    group = StudentGroup.query.get(target.group_id)
    if not group:
        raise ValueError(f"StudentGroup with ID {target.group_id} does not exist")
    # Verify teacher exists
    teacher = User.query.get(target.teacher_id)
    if not teacher:
        raise ValueError(f"Teacher with ID {target.teacher_id} does not exist")

# Database integrity check helper
def verify_data_integrity():
    """Perform comprehensive data integrity check"""
    issues = []
    
    # Check Remarks
    for remark in Remark.query.all():
        if not User.query.get(remark.teacher_id):
            issues.append(f"Remark {remark.id}: Invalid teacher_id {remark.teacher_id}")
        if not StudentGroup.query.get(remark.group_id):
            issues.append(f"Remark {remark.id}: Invalid group_id {remark.group_id}")
    
    # Check TeacherSchedule
    for ts in TeacherSchedule.query.all():
        if not User.query.get(ts.teacher_id):
            issues.append(f"TeacherSchedule {ts.id}: Invalid teacher_id {ts.teacher_id}")
        if not TimeSlot.query.get(ts.time_slot_id):
            issues.append(f"TeacherSchedule {ts.id}: Invalid time_slot_id {ts.time_slot_id}")
    
    # Check RoomSchedule
    for rs in RoomSchedule.query.all():
        if not Room.query.get(rs.room_id):
            issues.append(f"RoomSchedule {rs.id}: Invalid room_id {rs.room_id}")
        if not TimeSlot.query.get(rs.time_slot_id):
            issues.append(f"RoomSchedule {rs.id}: Invalid time_slot_id {rs.time_slot_id}")
    
    # Check StudentGroup
    for group in StudentGroup.query.all():
        if group.supervisor_id and not User.query.get(group.supervisor_id):
            issues.append(f"StudentGroup {group.id}: Invalid supervisor_id {group.supervisor_id}")
    
    # Check Viva
    for viva in Viva.query.all():
        if not StudentGroup.query.get(viva.group_id):
            issues.append(f"Viva {viva.id}: Invalid group_id {viva.group_id}")
        if not User.query.get(viva.teacher_id):
            issues.append(f"Viva {viva.id}: Invalid teacher_id {viva.teacher_id}")
    
    return issues

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    flash('Your session expired. Please try again.', 'warning')
    return redirect(request.referrer or url_for('dashboard'))

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        selected_role = request.form.get('role')
        remember = request.form.get('rememberMe') == 'on'
        
        # Get client IP and user agent for logging
        client_ip = request.remote_addr
        user_agent = request.headers.get('User-Agent', 'Unknown')[:255]
        
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            # Check if the selected role matches the user's role in the database
            # Treat 'teacher' and 'faculty' as interchangeable roles
            user_role_normalized = 'faculty' if user.role == 'teacher' else user.role
            selected_role_normalized = 'faculty' if selected_role == 'teacher' else selected_role
            
            if user_role_normalized != selected_role_normalized:
                # Log failed attempt due to role mismatch
                login_attempt = LoginAttempt(
                    email=email,
                    success=False,
                    ip_address=client_ip,
                    user_agent=user_agent,
                    user_id=user.id
                )
                db.session.add(login_attempt)
                db.session.commit()
                
                flash(f'Invalid role selected. You are registered as a {user.role}.', 'danger')
                return render_template('login_simple.html')
            
            # Log successful login
            login_attempt = LoginAttempt(
                email=email,
                success=True,
                ip_address=client_ip,
                user_agent=user_agent,
                user_id=user.id
            )
            db.session.add(login_attempt)
            db.session.commit()
            
            login_user(user, remember=remember)
            next_page = request.args.get('next')
            
            # Redirect user to their role-based dashboard
            return redirect(next_page or url_for('dashboard'))
        else:
            # Log failed login attempt
            login_attempt = LoginAttempt(
                email=email,
                success=False,
                ip_address=client_ip,
                user_agent=user_agent,
                user_id=user.id if user else None
            )
            db.session.add(login_attempt)
            db.session.commit()
            
            flash('Invalid email or password', 'danger')
    
    return render_template('login_simple.html')

@app.route('/login/modern')
def login_modern():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('login_modern.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirmPassword')
        first_name = request.form.get('firstName')
        last_name = request.form.get('lastName')
        role = request.form.get('role')
        username = request.form.get('username')
        
        # Get extended fields
        program = request.form.get('program')
        semester = request.form.get('semester')
        highest_degree = request.form.get('highestDegree')
        specialization = request.form.get('specialization')
        affiliation = request.form.get('affiliation')
        other_affiliation = request.form.get('otherAffiliation')
        
        # Validate role - only allow student and faculty, not supervisor/admin
        allowed_roles = ['student', 'faculty']
        if role not in allowed_roles:
            flash(f'Invalid role. Allowed roles: {", ".join(allowed_roles)}', 'danger')
            return render_template('signup.html')
        
        # Validate password
        if len(password) < 6:
            flash('Password must be at least 6 characters long', 'danger')
            return render_template('signup.html')
        
        if password != confirm_password:
            flash('Passwords do not match', 'danger')
            return render_template('signup.html')
        
        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            flash(f'Email already registered. You are registered as a {existing_user.role}. Please login.', 'danger')
            return render_template('signup.html')
            
        # Check faculty username if registering as faculty
        if role == 'faculty':
            if not username:
                flash('Username is required for faculty accounts', 'danger')
                return render_template('signup.html')
            
            teacher_username = TeacherUsername.query.filter_by(username=username.lower()).first()
            if not teacher_username:
                flash('Invalid faculty username. Please contact an administrator.', 'danger')
                return render_template('signup.html')
            
            if teacher_username.is_used:
                flash('This faculty username is already in use', 'danger')
                return render_template('signup.html')
        
        user = User(
            email=email,
            first_name=first_name,
            last_name=last_name,
            role=role,
            program=program,
            semester=semester,
            highest_degree=highest_degree,
            specialization=specialization,
            affiliation=affiliation,
            other_affiliation=other_affiliation
        )
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        # Mark faculty username as used if applicable
        if role == 'faculty' and username:
            teacher_username = TeacherUsername.query.filter_by(username=username.lower()).first()
            if teacher_username:
                teacher_username.is_used = True
                teacher_username.user_id = user.id
        db.session.commit()
        
        flash('Account created successfully! Please login.', 'success')
        return redirect(url_for('login'))
    
    return render_template('signup.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/profile')
@login_required
def profile():
    """Display user profile page"""
    return render_template('profile.html', user=current_user)

@app.route('/dashboard')
@login_required
def dashboard():
    """Generic dashboard route that redirects based on user role"""
    if current_user.role == 'admin':
        return redirect(url_for('dashboard_admin'))
    elif current_user.role in ('faculty', 'teacher'):
        return redirect(url_for('dashboard_faculty'))
    elif current_user.role == 'supervisor':
        return redirect(url_for('dashboard_supervisor'))
    elif current_user.role == 'student':
        return redirect(url_for('dashboard_student'))
    else:
        return redirect(url_for('index'))

# Role-based access decorator
def role_required(role):
    def decorator(f):
        @wraps(f)
        @login_required
        def decorated_function(*args, **kwargs):
            if current_user.role != role:
                flash('You do not have access to this dashboard.', 'danger')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.route('/delete_user/<int:user_id>')
@login_required
def delete_user(user_id):
    if current_user.role != 'admin':
        flash('You do not have permission to delete users.', 'danger')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(user_id)
    
    db.session.delete(user)
    db.session.commit()
    
    flash(f'User {user.email} has been deleted.', 'success')
    return redirect(url_for('dashboard'))

@app.route('/dashboard_faculty')
@login_required
def dashboard_faculty():
    # Allow faculty and teacher roles (teacher is legacy name for faculty)
    allowed_roles = {"faculty", "teacher"}
    if current_user.role not in allowed_roles:
        flash("Access denied.")
        return redirect(url_for('dashboard'))
    
    # Get all groups supervised by the current user
    supervised_groups = StudentGroup.query.filter_by(supervisor_id=current_user.id).all()
    
    # Get all students in the supervised groups
    supervised_student_ids = []
    for group in supervised_groups:
        group_members = GroupMember.query.filter_by(group_id=group.id).all()
        supervised_student_ids.extend([member.user_id for member in group_members])
    
    # Fetch the actual student objects
    supervised_students = User.query.filter(User.id.in_(supervised_student_ids)).all() if supervised_student_ids else []
    
    # Get all remarks for the supervised groups
    all_remarks = []
    for group in supervised_groups:
        remarks = Remark.query.filter_by(group_id=group.id).order_by(Remark.timestamp.desc()).all()
        for remark in remarks:
            all_remarks.append({
                'content': remark.content,
                'teacher_name': f"{remark.teacher.first_name} {remark.teacher.last_name}",
                'timestamp': remark.timestamp,
                'group_id': group.group_id,
                'project_title': group.project_title
            })
    
    # Get today's and tomorrow's schedules
    today = datetime.datetime.now().strftime('%A')
    tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%A')
    
    # Get all room schedules (all classes)
    room_schedules = RoomSchedule.query.all()
    
    # Get all time slots and rooms for reference
    time_slots = {ts.id: ts for ts in TimeSlot.query.all()}
    rooms = {r.id: r for r in Room.query.all()}
    
    # Get teacher information for each schedule
    teacher_schedules = TeacherSchedule.query.all()
    teacher_map = {}
    for ts in teacher_schedules:
        key = f"{ts.time_slot_id}"
        teacher = User.query.get(ts.teacher_id)
        if teacher:
            teacher_map[key] = {
                'name': f"{teacher.first_name} {teacher.last_name}",
                'subject': ts.subject
            }
    
    today_schedules = []
    tomorrow_schedules = []
    
    for schedule in room_schedules:
        if schedule.time_slot_id in time_slots and schedule.room_id in rooms:
            time_slot = time_slots[schedule.time_slot_id]
            room = rooms[schedule.room_id]
            
            # Get teacher info if available
            teacher_info = teacher_map.get(f"{schedule.time_slot_id}", None)
            teacher_name = teacher_info['name'] if teacher_info else "No teacher assigned"
            subject = teacher_info['subject'] if teacher_info else schedule.class_name
            
            schedule_info = {
                'id': schedule.id,
                'time': f"{time_slot.start_time} - {time_slot.end_time}",
                'start_time': time_slot.start_time,
                'class_name': schedule.class_name,
                'room': room.name,
                'teacher': teacher_name,
                'subject': subject,
                'day': time_slot.day
            }
            
            if time_slot.day == today:
                today_schedules.append(schedule_info)
            elif time_slot.day == tomorrow:
                tomorrow_schedules.append(schedule_info)
    
    # Sort schedules by start time
    today_schedules.sort(key=lambda x: x['start_time'])
    tomorrow_schedules.sort(key=lambda x: x['start_time'])
    
    # Get assigned works for all supervised groups
    supervised_group_ids = [g.id for g in supervised_groups]
    all_assigned_works = AssignedWork.query.filter(
        AssignedWork.group_id.in_(supervised_group_ids)
    ).order_by(
        db.case((AssignedWork.due_date == None, 1), else_=0),
        AssignedWork.due_date.asc()
    ).all() if supervised_group_ids else []
    
    # Get vivas for supervised groups
    all_vivas = Viva.query.filter(
        Viva.group_id.in_(supervised_group_ids)
    ).order_by(Viva.scheduled_date.asc()).all() if supervised_group_ids else []
    
    # Get project statuses for supervised groups
    all_project_statuses = {ps.group_id: ps for ps in ProjectStatus.query.filter(
        ProjectStatus.group_id.in_(supervised_group_ids)
    ).all()} if supervised_group_ids else {}
    
    # Get project details (progress) for supervised groups
    all_project_details = {pd.group_id: pd for pd in ProjectDetails.query.filter(
        ProjectDetails.group_id.in_(supervised_group_ids)
    ).all()} if supervised_group_ids else {}
    
    # Calculate overall project progress from real data
    if all_project_details:
        total_progress = sum(pd.progress for pd in all_project_details.values() if pd.progress)
        project_progress = total_progress // len(all_project_details) if all_project_details else 0
    elif all_assigned_works:
        completed = sum(1 for w in all_assigned_works if w.status == 'Completed')
        project_progress = int((completed / len(all_assigned_works)) * 100) if all_assigned_works else 0
    else:
        project_progress = 0
    
    # Get groups assigned to this faculty for evaluation (via ProjectStatus)
    assigned_project_statuses = ProjectStatus.query.filter_by(teacher_id=current_user.id).all()
    assigned_group_ids = [ps.group_id for ps in assigned_project_statuses]
    assigned_groups = StudentGroup.query.filter(StudentGroup.id.in_(assigned_group_ids)).all() if assigned_group_ids else []
    
    # Get count of assigned groups
    assigned_groups_count = len(assigned_groups)
    pending_evaluations_count = ProjectStatus.query.filter_by(teacher_id=current_user.id, status='Pending').count()
    
    # Get submissions for supervised groups
    all_submissions = Submission.query.filter(Submission.group_id.in_(supervised_group_ids)).order_by(Submission.created_at.desc()).all() if supervised_group_ids else []
    
    return render_template('dashboard_faculty.html', 
                          current_user=current_user,
                          today_schedules=today_schedules,
                          tomorrow_schedules=tomorrow_schedules,
                          groups=supervised_groups,
                          assigned_groups=assigned_groups,
                          students=supervised_students,
                          remarks=all_remarks,
                          today=today,
                          assigned_groups_count=assigned_groups_count,
                          pending_evaluations_count=pending_evaluations_count,
                          all_assigned_works=all_assigned_works,
                          all_vivas=all_vivas,
                          all_project_statuses=all_project_statuses,
                          all_project_details=all_project_details,
                          project_progress=project_progress,
                          all_submissions=all_submissions)

@app.route('/dashboard_admin')
@login_required
def dashboard_admin():
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    # Get query parameters
    page = request.args.get('page', 1, type=int)
    per_page = 20  # Display 20 users per page
    search_query = request.args.get('search', '')
    role_filter = request.args.get('role', '')
    
    # Start with base query
    query = User.query
    
    # Apply search filter if provided
    if search_query:
        search_term = f"%{search_query}%"
        query = query.filter(
            db.or_(
                User.email.like(search_term),
                User.first_name.like(search_term),
                User.last_name.like(search_term)
            )
        )
    
    # Apply role filter if provided
    if role_filter and role_filter != 'all':
        query = query.filter(User.role == role_filter)
    
    # Paginate the results
    users_paginated = query.paginate(page=page, per_page=per_page, error_out=False)
    users = users_paginated.items
    
    # Get all users for counting
    all_users = User.query.all()
    
    # Count users by role (treating 'teacher' and 'faculty' as the same)
    all_students = [u for u in all_users if u.role == 'student']
    teachers = [u for u in all_users if u.role in ('teacher', 'faculty')]
    supervisors = [u for u in all_users if u.role == 'supervisor']
    admins = [u for u in all_users if u.role == 'admin']
    
    # Get only available students (not already assigned to any group)
    assigned_student_ids = db.session.query(GroupMember.user_id).all()
    assigned_ids = [student[0] for student in assigned_student_ids]
    students = [u for u in all_students if u.id not in assigned_ids]
    
    # Get all projects/groups
    projects = StudentGroup.query.all()
    
    # Calculate project status counts based on actual ProjectStatus records
    # Get the latest status for each group
    latest_statuses = {}
    all_project_statuses = ProjectStatus.query.order_by(ProjectStatus.created_at.desc()).all()
    
    for ps in all_project_statuses:
        if ps.group_id not in latest_statuses:
            latest_statuses[ps.group_id] = ps.status
    
    # Get all groups
    groups = StudentGroup.query.all()
    
    # Count groups with no status record as 'Pending'
    groups_without_status = len([g for g in groups if g.id not in latest_statuses])
    
    pending_count = list(latest_statuses.values()).count('Pending') + groups_without_status
    accepted_count = list(latest_statuses.values()).count('Accepted')
    conditional_count = list(latest_statuses.values()).count('Conditionally Accepted')
    deferred_count = list(latest_statuses.values()).count('Deferred')
    completed = list(latest_statuses.values()).count('Completed')
    
    # Keep old variable names for stat cards compatibility
    on_track = accepted_count
    at_risk = conditional_count
    delayed = deferred_count
    
    # Get all remarks
    remarks = Remark.query.all()
    
    # Calculate system usage analytics
    from datetime import datetime, timedelta
    now = datetime.now()

    # Last 7 days and last 30 days
    last_7_days = now - timedelta(days=7)
    last_30_days = now - timedelta(days=30)

    # New users in last 7 and 30 days (users created in that period)
    new_users_7_days = User.query.filter(User.created_at >= last_7_days).count()
    new_users_30_days = User.query.filter(User.created_at >= last_30_days).count()

    # For now, use remark-based activity as proxy since we don't have user creation timestamps
    # Project activity (count actual remarks as activities)
    project_activity_7_days = len([r for r in remarks if r.timestamp and r.timestamp > last_7_days]) if remarks else 0
    project_activity_30_days = len([r for r in remarks if r.timestamp and r.timestamp > last_30_days]) if remarks else 0

    # Remarks added
    remarks_7_days = project_activity_7_days
    remarks_30_days = project_activity_30_days

    # Login activity - using actual login attempt data
    logins_7_days = LoginAttempt.query.filter(
        LoginAttempt.success == True,
        LoginAttempt.timestamp >= last_7_days
    ).count()
    logins_30_days = LoginAttempt.query.filter(
        LoginAttempt.success == True,
        LoginAttempt.timestamp >= last_30_days
    ).count()
    
    # Failed login attempts
    failed_logins_7_days = LoginAttempt.query.filter(
        LoginAttempt.success == False,
        LoginAttempt.timestamp >= last_7_days
    ).count()
    failed_logins_30_days = LoginAttempt.query.filter(
        LoginAttempt.success == False,
        LoginAttempt.timestamp >= last_30_days
    ).count()
    
    # Recent login activity
    recent_logins = LoginAttempt.query.order_by(LoginAttempt.timestamp.desc()).limit(10).all()

    # Build recent activity feed from real data
    import datetime as dt
    now = dt.datetime.utcnow()
    recent_activities = []
    
    # Recent users
    recent_users = User.query.order_by(User.created_at.desc()).limit(5).all()
    for u in recent_users:
        if u.created_at:
            recent_activities.append({
                'icon': 'bi-person-plus',
                'icon_class': 'success',
                'title': f'New {u.role} registered: {u.first_name} {u.last_name}',
                'timestamp': u.created_at
            })
    
    # Recent groups created
    recent_groups_list = StudentGroup.query.order_by(StudentGroup.created_at.desc()).limit(5).all()
    for g in recent_groups_list:
        if g.created_at:
            recent_activities.append({
                'icon': 'bi-folder-plus',
                'icon_class': 'info',
                'title': f'Group created: {g.group_id} - {g.project_title}',
                'timestamp': g.created_at
            })
    
    # Recent vivas scheduled
    recent_vivas = Viva.query.order_by(Viva.created_at.desc()).limit(5).all()
    for v in recent_vivas:
        if v.created_at:
            recent_activities.append({
                'icon': 'bi-calendar-check',
                'icon_class': 'info',
                'title': f'Viva scheduled for {v.group.group_id} on {v.scheduled_date.strftime("%d-%m-%Y")}',
                'timestamp': v.created_at
            })
    
    # Recent remarks
    recent_remarks_list = Remark.query.order_by(Remark.timestamp.desc()).limit(5).all()
    for r in recent_remarks_list:
        if r.timestamp:
            teacher = User.query.get(r.teacher_id)
            group = StudentGroup.query.get(r.group_id)
            teacher_name = f'{teacher.first_name} {teacher.last_name}' if teacher else 'Unknown'
            group_name = group.group_id if group else 'Unknown'
            recent_activities.append({
                'icon': 'bi-chat-left-text',
                'icon_class': 'warning',
                'title': f'Remark by {teacher_name} on {group_name}',
                'timestamp': r.timestamp
            })
    
    # Recent failed logins
    recent_failed = LoginAttempt.query.filter_by(success=False).order_by(LoginAttempt.timestamp.desc()).limit(3).all()
    for lf in recent_failed:
        if lf.timestamp:
            recent_activities.append({
                'icon': 'bi-exclamation-triangle',
                'icon_class': 'warning',
                'title': f'Failed login attempt: {lf.email}',
                'timestamp': lf.timestamp
            })
    
    # Sort all by timestamp descending & take top 5
    recent_activities.sort(key=lambda x: x['timestamp'], reverse=True)
    recent_activities = recent_activities[:5]
    
    # Calculate time ago for each activity
    def time_ago(ts):
        if not ts:
            return 'Unknown'
        diff = now - ts
        seconds = diff.total_seconds()
        if seconds < 60:
            return 'Just now'
        elif seconds < 3600:
            mins = int(seconds // 60)
            return f'{mins} min{"s" if mins > 1 else ""} ago'
        elif seconds < 86400:
            hours = int(seconds // 3600)
            return f'{hours} hour{"s" if hours > 1 else ""} ago'
        elif seconds < 604800:
            days = int(seconds // 86400)
            return f'{days} day{"s" if days > 1 else ""} ago'
        else:
            return ts.strftime('%d %b %Y')
    
    for a in recent_activities:
        a['time_ago'] = time_ago(a['timestamp'])
    
    # System status - real checks
    try:
        db.session.execute(db.text('SELECT 1'))
        db_status = 'Connected'
        db_ok = True
    except Exception:
        db_status = 'Error'
        db_ok = False
    
    system_status = {
        'db_status': db_status,
        'db_ok': db_ok,
        'total_users': len(all_users),
        'total_groups': len(projects),
        'total_vivas': Viva.query.count(),
        'pending_vivas': Viva.query.filter_by(status='Scheduled').count(),
    }

    # Trends calculation: compare last 7 days vs the 7 days before that (days 8-14)
    def calculate_trend(current_period, previous_period):
        if previous_period == 0:
            return 100 if current_period > 0 else 0
        return int(((current_period - previous_period) / previous_period) * 100)

    last_14_days = now - timedelta(days=14)
    
    # Login trends (week over week)
    logins_prev_week = LoginAttempt.query.filter(
        LoginAttempt.success == True,
        LoginAttempt.timestamp >= last_14_days,
        LoginAttempt.timestamp < last_7_days
    ).count()
    login_trend = calculate_trend(logins_7_days, logins_prev_week)
    
    # User trends
    new_users_prev_week = User.query.filter(
        User.created_at >= last_14_days,
        User.created_at < last_7_days
    ).count()
    user_trend = calculate_trend(new_users_7_days, new_users_prev_week)
    
    # Project activity trends
    project_activity_prev_week = len([r for r in remarks if r.timestamp and r.timestamp >= last_14_days and r.timestamp < last_7_days]) if remarks else 0
    project_trend = calculate_trend(project_activity_7_days, project_activity_prev_week)
    
    # Remarks trends
    remarks_prev_week = project_activity_prev_week
    remarks_trend = calculate_trend(remarks_7_days, remarks_prev_week)
    
    return render_template('dashboard_admin_modern.html', 
                          total_users=len(all_users),
                          total_projects=len(projects),
                          total_groups=len(projects),
                          users=users,
                          pagination=users_paginated,
                          students=students,
                          teachers=teachers,
                          supervisors=supervisors,
                          admins=admins,
                          projects=projects,
                          groups=groups,
                          remarks=remarks,
                          search_query=search_query,
                          role_filter=role_filter,
                          on_track=on_track,
                          at_risk=at_risk,
                          delayed=delayed,
                          completed=completed,
                          pending_count=pending_count,
                          accepted_count=accepted_count,
                          conditional_count=conditional_count,
                          deferred_count=deferred_count,
                          # System usage analytics data
                          logins_7_days=logins_7_days,
                          logins_30_days=logins_30_days, 
                          new_users_7_days=new_users_7_days,
                          new_users_30_days=new_users_30_days,
                          project_activity_7_days=project_activity_7_days,
                          project_activity_30_days=project_activity_30_days,
                          remarks_7_days=remarks_7_days,
                          remarks_30_days=remarks_30_days,
                          login_trend=login_trend,
                          user_trend=user_trend,
                          project_trend=project_trend,
                          remarks_trend=remarks_trend,
                          failed_logins_7_days=failed_logins_7_days,
                          failed_logins_30_days=failed_logins_30_days,
                          recent_logins=recent_logins,
                          recent_activities=recent_activities,
                          system_status=system_status)

@app.route('/dashboard_supervisor')
@login_required
def dashboard_supervisor():
    if current_user.role != "supervisor":
        flash("Access denied.")
        return redirect(url_for('dashboard'))
    
    # Get only this supervisor's groups
    student_groups = StudentGroup.query.filter_by(supervisor_id=current_user.id).all()
    
    # Get all remarks from teachers for their groups
    all_remarks = []
    for group in student_groups:
        remarks = Remark.query.filter_by(group_id=group.id).order_by(Remark.timestamp.desc()).all()
        for remark in remarks:
            teacher = User.query.get(remark.teacher_id)
            student_name = None
            if remark.student_id:
                student = User.query.get(remark.student_id)
                if student:
                    student_name = f"{student.first_name} {student.last_name}"
            if teacher:
                all_remarks.append({
                    'content': remark.content,
                    'teacher_name': f"{teacher.first_name} {teacher.last_name}",
                    'timestamp': remark.timestamp,
                    'group_id': group.group_id,
                    'project_title': group.project_title,
                    'student_name': student_name
                })
    
    # Get milestones for all supervised groups (kept for faculty)
    group_ids = [g.id for g in student_groups]
    
    # Get vivas for supervised groups
    all_vivas = Viva.query.filter(Viva.group_id.in_(group_ids)).order_by(Viva.scheduled_date.asc()).all() if group_ids else []
    
    # Get project statuses
    all_project_statuses = {ps.group_id: ps for ps in ProjectStatus.query.filter(ProjectStatus.group_id.in_(group_ids)).all()} if group_ids else {}
    
    # Get project details (progress)
    all_project_details = {pd.group_id: pd for pd in ProjectDetails.query.filter(ProjectDetails.group_id.in_(group_ids)).all()} if group_ids else {}
    
    # Get submissions for supervised groups
    all_submissions = Submission.query.filter(Submission.group_id.in_(group_ids)).order_by(Submission.created_at.desc()).all() if group_ids else []
    
    # Get assigned works for supervised groups
    all_assigned_works = AssignedWork.query.filter(AssignedWork.group_id.in_(group_ids)).order_by(AssignedWork.created_at.desc()).all() if group_ids else []
    
    # Auto-mark overdue works
    for work in all_assigned_works:
        if work.is_overdue and work.status == 'Pending':
            work.status = 'Overdue'
    db.session.commit()

    # Build a dict of group members for the assign-work form
    group_members = {}
    for g in student_groups:
        members = GroupMember.query.filter_by(group_id=g.id).all()
        group_members[g.id] = [{'id': m.user_id, 'name': f"{m.user.first_name} {m.user.last_name}"} for m in members]

    return render_template('dashboard_supervisor.html', 
                           supervised_groups=student_groups,
                           all_remarks=all_remarks,
                           all_vivas=all_vivas,
                           all_project_statuses=all_project_statuses,
                           all_project_details=all_project_details,
                           all_submissions=all_submissions,
                           all_assigned_works=all_assigned_works,
                           group_members=group_members)

@app.route('/dashboard/student')
@role_required('student')
def dashboard_student():
    # Get today's and tomorrow's day names
    today = datetime.datetime.now().strftime('%A')
    tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime('%A')
    
    # Get the student's group if they're assigned to one
    student_group_membership = GroupMember.query.filter_by(user_id=current_user.id).first()
    student_group = None
    
    if student_group_membership:
        student_group = StudentGroup.query.get(student_group_membership.group_id)
    
    # Get the supervisor for the group
    supervisor = None
    if student_group and student_group.supervisor_id:
        supervisor = User.query.get(student_group.supervisor_id)
    
    # Get remarks/feedback for the group
    remarks = []
    if student_group:
        # Show remarks for the whole group (student_id is NULL) or specifically for this student
        remarks = Remark.query.filter_by(group_id=student_group.id).filter(
            db.or_(Remark.student_id == None, Remark.student_id == current_user.id)
        ).order_by(Remark.timestamp.desc()).all()
    
    # Get all room schedules (all classes)
    room_schedules = RoomSchedule.query.all()
    
    # Get all time slots and rooms for reference
    time_slots = {ts.id: ts for ts in TimeSlot.query.all()}
    rooms = {r.id: r for r in Room.query.all()}
    
    # Get teacher information for each schedule
    teacher_schedules = TeacherSchedule.query.all()
    teacher_map = {}
    for ts in teacher_schedules:
        key = f"{ts.time_slot_id}"
        teacher = User.query.get(ts.teacher_id)
        if teacher:
            teacher_map[key] = {
                'name': f"{teacher.first_name} {teacher.last_name}",
                'subject': ts.subject
            }
    
    # Organize schedules by day for easier rendering
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    schedules_by_day = {day: [] for day in days}
    
    for schedule in room_schedules:
        if schedule.time_slot_id in time_slots and schedule.room_id in rooms:
            time_slot = time_slots[schedule.time_slot_id]
            room = rooms[schedule.room_id]
            
            # Get teacher info if available
            teacher_info = teacher_map.get(f"{schedule.time_slot_id}", None)
            teacher_name = teacher_info['name'] if teacher_info else "No teacher assigned"
            subject = teacher_info['subject'] if teacher_info else schedule.class_name
            
            schedules_by_day[time_slot.day].append({
                'id': schedule.id,
                'time': f"{time_slot.start_time} - {time_slot.end_time}",
                'class_name': schedule.class_name,
                'room': room.name,
                'teacher': teacher_name,
                'subject': subject
            })
    
    # Sort each day's schedules by time
    for day in days:
        schedules_by_day[day].sort(key=lambda x: x['time'])
    
    # Get progress for student's group
    project_progress = 0
    if student_group:
        # Get real progress from ProjectDetails
        details = ProjectDetails.query.filter_by(group_id=student_group.id).first()
        if details:
            project_progress = details.progress
    
    # Get student's submissions
    submissions = []
    if student_group:
        submissions = Submission.query.filter_by(group_id=student_group.id, student_id=current_user.id).order_by(Submission.created_at.desc()).all()

    # Get assigned works for this student
    assigned_works = []
    if student_group:
        assigned_works = AssignedWork.query.filter(
            AssignedWork.group_id == student_group.id,
            db.or_(AssignedWork.assigned_to == current_user.id, AssignedWork.assigned_to == None)
        ).order_by(db.case((AssignedWork.due_date == None, 1), else_=0), AssignedWork.due_date.asc(), AssignedWork.created_at.desc()).all()
        # Auto-mark overdue
        for w in assigned_works:
            if w.is_overdue and w.status == 'Pending':
                w.status = 'Overdue'
        db.session.commit()

    return render_template(
        'dashboard_student.html',
        student_group=student_group,
        supervisor=supervisor,
        remarks=remarks,
        today_schedules=schedules_by_day.get(today, []),
        tomorrow_schedules=schedules_by_day.get(tomorrow, []),
        today=today,
        tomorrow=tomorrow,
        project_progress=project_progress,
        submissions=submissions,
        assigned_works=assigned_works,
        current_user=current_user
    )

@app.route('/admin/db')
@role_required('admin')
def admin_db():
    users = User.query.all()
    return render_template('admin_db.html', users=users)

@app.route('/add_remark', methods=['POST'])
@login_required
def add_remark():
    if current_user.role not in ("faculty", "teacher", "admin", "supervisor"):
        flash("Access denied.")
        return redirect(url_for('index'))
    
    group_id = request.form.get('group_id')
    student_id = request.form.get('student_id')  # Optional: specific student
    content = request.form.get('content')
    
    try:
        group = StudentGroup.query.filter_by(id=group_id).first()
        if not group:
            flash("Group not found", 'danger')
            return redirect(url_for('dashboard'))
        
        # Supervisors can only remark on their own groups
        if current_user.role == 'supervisor' and group.supervisor_id != current_user.id:
            flash("You can only send remarks to your own groups", 'danger')
            return redirect(url_for('dashboard'))
        
        if not content or len(str(content).strip()) == 0:
            flash("Remark content cannot be empty", 'danger')
            return redirect(url_for('dashboard'))
        
        # Validate student_id if provided
        target_student_id = None
        if student_id and student_id.strip():
            target_student_id = int(student_id)
            # Verify this student is a member of the selected group
            member = GroupMember.query.filter_by(user_id=target_student_id, group_id=group.id).first()
            if not member:
                flash("Selected student is not a member of this group", 'danger')
                return redirect(url_for('dashboard'))
        
        new_remark = Remark(
            content=content,
            teacher_id=current_user.id,
            group_id=group.id,
            student_id=target_student_id
        )
        
        db.session.add(new_remark)
        db.session.commit()
        
        if target_student_id:
            student = User.query.get(target_student_id)
            logger.info(f"Remark added by {current_user.email} for student {student.email} in group {group.group_id}")
        else:
            logger.info(f"Remark added by {current_user.email} for group {group.group_id}")
        flash("Remark added successfully", 'success')
        return redirect(url_for('dashboard'))
    
    except ValueError as ve:
        db.session.rollback()
        logger.error(f"Validation error adding remark: {str(ve)}")
        flash(f"Error adding remark: {str(ve)}", 'danger')
        return redirect(url_for('dashboard'))
    except Exception as e:
        db.session.rollback()
        logger.error(f"Unexpected error adding remark: {str(e)}")
        flash(f"Unexpected error: {str(e)}", 'danger')
        return redirect(url_for('dashboard'))

# ========== SUBMISSION ROUTES ==========

@app.route('/student/submit_work', methods=['POST'])
@login_required
def student_submit_work():
    if current_user.role != 'student':
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))
    
    # Check if student is in a group
    membership = GroupMember.query.filter_by(user_id=current_user.id).first()
    if not membership:
        flash("You must be assigned to a group before submitting work.", 'warning')
        return redirect(url_for('dashboard'))
    
    title = request.form.get('submission_title', '').strip()
    description = request.form.get('submission_description', '').strip()
    submission_type = request.form.get('submission_type', 'General')
    file = request.files.get('submission_file')
    
    if not title:
        flash("Submission title is required.", 'danger')
        return redirect(url_for('dashboard'))
    
    if not file or file.filename == '':
        flash("Please select a file to upload.", 'danger')
        return redirect(url_for('dashboard'))
    
    if not allowed_file(file.filename):
        flash(f"File type not allowed. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}", 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        original_filename = secure_filename(file.filename)
        # Create unique filename to avoid conflicts
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        stored_filename = f"{current_user.id}_{membership.group_id}_{timestamp}_{original_filename}"
        
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], stored_filename)
        file.save(filepath)
        file_size = os.path.getsize(filepath)
        file_ext = original_filename.rsplit('.', 1)[1].lower() if '.' in original_filename else ''
        
        submission = Submission(
            title=title,
            description=description,
            filename=stored_filename,
            original_filename=original_filename,
            file_size=file_size,
            file_type=file_ext,
            submission_type=submission_type,
            student_id=current_user.id,
            group_id=membership.group_id
        )
        db.session.add(submission)
        db.session.commit()
        
        flash(f"'{title}' submitted successfully!", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Error uploading file: {str(e)}", 'danger')
    
    return redirect(url_for('dashboard'))


@app.route('/submission/download/<int:submission_id>')
@login_required
def download_submission(submission_id):
    submission = Submission.query.get_or_404(submission_id)
    
    # Access control: student (own group), supervisor (own groups), faculty/teacher, admin
    if current_user.role == 'student':
        membership = GroupMember.query.filter_by(user_id=current_user.id, group_id=submission.group_id).first()
        if not membership:
            flash("Access denied.", 'danger')
            return redirect(url_for('dashboard'))
    elif current_user.role == 'supervisor':
        group = StudentGroup.query.get(submission.group_id)
        if not group or group.supervisor_id != current_user.id:
            flash("Access denied.", 'danger')
            return redirect(url_for('dashboard'))
    elif current_user.role not in ('faculty', 'teacher', 'admin'):
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))
    
    return send_from_directory(
        app.config['UPLOAD_FOLDER'],
        submission.filename,
        as_attachment=True,
        download_name=submission.original_filename
    )


@app.route('/student/delete_submission/<int:submission_id>', methods=['POST'])
@login_required
def delete_submission(submission_id):
    if current_user.role != 'student':
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))
    
    submission = Submission.query.get_or_404(submission_id)
    
    if submission.student_id != current_user.id:
        flash("You can only delete your own submissions.", 'danger')
        return redirect(url_for('dashboard'))
    
    if submission.status in ('Approved', 'Reviewed'):
        flash("Cannot delete a submission that has already been reviewed.", 'warning')
        return redirect(url_for('dashboard'))
    
    try:
        # Delete the file from disk
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], submission.filename)
        if os.path.exists(filepath):
            os.remove(filepath)
        
        db.session.delete(submission)
        db.session.commit()
        flash("Submission deleted successfully.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting submission: {str(e)}", 'danger')
    
    return redirect(url_for('dashboard'))


@app.route('/submission/review/<int:submission_id>', methods=['POST'])
@login_required
def review_submission(submission_id):
    if current_user.role not in ('supervisor', 'faculty', 'teacher', 'admin'):
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))
    
    submission = Submission.query.get_or_404(submission_id)
    
    # Supervisor can only review their own groups' submissions
    if current_user.role == 'supervisor':
        group = StudentGroup.query.get(submission.group_id)
        if not group or group.supervisor_id != current_user.id:
            flash("You can only review submissions from your own groups.", 'danger')
            return redirect(url_for('dashboard'))
    
    new_status = request.form.get('status', 'Reviewed')
    feedback = request.form.get('feedback', '').strip()
    
    if new_status not in ('Reviewed', 'Approved', 'Rejected'):
        flash("Invalid status.", 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        submission.status = new_status
        submission.feedback = feedback
        db.session.commit()
        flash(f"Submission marked as '{new_status}'.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Error updating submission: {str(e)}", 'danger')
    
    return redirect(url_for('dashboard'))

# ========== ASSIGN WORK ROUTES ==========

@app.route('/supervisor/assign_work', methods=['POST'])
@login_required
def assign_work():
    if current_user.role != 'supervisor':
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))

    group_id = request.form.get('work_group_id', type=int)
    title = request.form.get('work_title', '').strip()
    description = request.form.get('work_description', '').strip()
    due_date_str = request.form.get('work_due_date', '').strip()
    priority = request.form.get('work_priority', 'Medium')
    work_type = request.form.get('work_type', 'General')
    assigned_to_val = request.form.get('work_assigned_to', '').strip()  # '' = whole group, or user_id

    if not group_id or not title:
        flash("Group and title are required.", 'danger')
        return redirect(url_for('dashboard'))

    # Verify supervisor owns this group
    group = StudentGroup.query.get(group_id)
    if not group or group.supervisor_id != current_user.id:
        flash("You can only assign work to your own groups.", 'danger')
        return redirect(url_for('dashboard'))

    due_date = None
    if due_date_str:
        try:
            due_date = datetime.datetime.strptime(due_date_str, '%Y-%m-%d').date()
        except ValueError:
            flash("Invalid date format.", 'danger')
            return redirect(url_for('dashboard'))

    assigned_to_id = None
    if assigned_to_val:
        assigned_to_id = int(assigned_to_val)
        # Verify the student is actually in the group
        member = GroupMember.query.filter_by(user_id=assigned_to_id, group_id=group_id).first()
        if not member:
            flash("Selected student is not in this group.", 'danger')
            return redirect(url_for('dashboard'))

    if priority not in ('Low', 'Medium', 'High', 'Urgent'):
        priority = 'Medium'

    try:
        work = AssignedWork(
            title=title,
            description=description,
            due_date=due_date,
            priority=priority,
            work_type=work_type,
            group_id=group_id,
            assigned_to=assigned_to_id,
            assigned_by=current_user.id
        )
        db.session.add(work)
        db.session.commit()

        student_name = "Entire Group"
        if assigned_to_id:
            student = User.query.get(assigned_to_id)
            student_name = f"{student.first_name} {student.last_name}" if student else "Student"

        flash(f"Work '{title}' assigned to {student_name} successfully!", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Error assigning work: {str(e)}", 'danger')

    return redirect(url_for('dashboard'))


@app.route('/supervisor/edit_work/<int:work_id>', methods=['POST'])
@login_required
def edit_assigned_work(work_id):
    if current_user.role != 'supervisor':
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))

    work = AssignedWork.query.get_or_404(work_id)
    group = StudentGroup.query.get(work.group_id)
    if not group or group.supervisor_id != current_user.id:
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))

    new_status = request.form.get('edit_work_status', '').strip()
    new_title = request.form.get('edit_work_title', '').strip()
    new_description = request.form.get('edit_work_description', '').strip()
    new_due_date_str = request.form.get('edit_work_due_date', '').strip()
    new_priority = request.form.get('edit_work_priority', '').strip()

    try:
        if new_title:
            work.title = new_title
        if new_description is not None:
            work.description = new_description
        if new_status and new_status in ('Pending', 'In Progress', 'Submitted', 'Completed', 'Overdue', 'Needs Revision'):
            work.status = new_status
        if new_priority and new_priority in ('Low', 'Medium', 'High', 'Urgent'):
            work.priority = new_priority
        if new_due_date_str:
            work.due_date = datetime.datetime.strptime(new_due_date_str, '%Y-%m-%d').date()

        db.session.commit()
        flash(f"Work '{work.title}' updated.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Error updating work: {str(e)}", 'danger')

    return redirect(url_for('dashboard'))


@app.route('/supervisor/delete_work/<int:work_id>', methods=['POST'])
@login_required
def delete_assigned_work(work_id):
    if current_user.role != 'supervisor':
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))

    work = AssignedWork.query.get_or_404(work_id)
    group = StudentGroup.query.get(work.group_id)
    if not group or group.supervisor_id != current_user.id:
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))

    try:
        title = work.title
        db.session.delete(work)
        db.session.commit()
        flash(f"Work '{title}' deleted.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting work: {str(e)}", 'danger')

    return redirect(url_for('dashboard'))


@app.route('/student/update_work/<int:work_id>', methods=['POST'])
@login_required
def student_update_work(work_id):
    if current_user.role != 'student':
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))

    work = AssignedWork.query.get_or_404(work_id)

    # Verify: student must be in the group AND either assigned_to is this student or the whole group
    membership = GroupMember.query.filter_by(user_id=current_user.id, group_id=work.group_id).first()
    if not membership:
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))
    if work.assigned_to and work.assigned_to != current_user.id:
        flash("This work is not assigned to you.", 'danger')
        return redirect(url_for('dashboard'))

    new_status = request.form.get('student_work_status', '').strip()
    response_text = request.form.get('student_work_response', '').strip()

    if new_status not in ('In Progress', 'Submitted'):
        flash("Invalid status update.", 'danger')
        return redirect(url_for('dashboard'))

    try:
        work.status = new_status
        if response_text:
            work.student_response = response_text
        db.session.commit()
        flash(f"Work '{work.title}' marked as '{new_status}'.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Error updating work: {str(e)}", 'danger')

    return redirect(url_for('dashboard'))


@app.route('/supervisor/review_work/<int:work_id>', methods=['POST'])
@login_required
def supervisor_review_work(work_id):
    if current_user.role != 'supervisor':
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))

    work = AssignedWork.query.get_or_404(work_id)
    group = StudentGroup.query.get(work.group_id)
    if not group or group.supervisor_id != current_user.id:
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))

    action = request.form.get('review_action', '').strip()
    comment = request.form.get('supervisor_comment', '').strip()

    try:
        if action == 'approve':
            work.status = 'Completed'
            flash_msg = f"Work '{work.title}' approved and marked as completed."
        elif action == 'needs_revision':
            work.status = 'Needs Revision'
            flash_msg = f"Work '{work.title}' sent back for revision."
        elif action == 'reject':
            work.status = 'Pending'
            flash_msg = f"Work '{work.title}' rejected and reset to Pending."
        else:
            flash("Invalid review action.", 'danger')
            return redirect(url_for('dashboard'))

        if comment:
            work.supervisor_comment = comment

        db.session.commit()
        logger.info(f"Supervisor {current_user.email} reviewed work '{work.title}': {action}")
        flash(flash_msg, 'success')
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error reviewing work: {str(e)}")
        flash(f"Error: {str(e)}", 'danger')

    return redirect(url_for('dashboard'))


# ========== EVALUATION ROUTES ==========

@app.route('/supervisor/update_progress', methods=['POST'])
@login_required
def supervisor_update_progress():
    if current_user.role != 'supervisor':
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))
    
    group_id = request.form.get('group_id')
    progress = request.form.get('progress', 0, type=int)
    
    try:
        group = StudentGroup.query.get(group_id)
        if not group or group.supervisor_id != current_user.id:
            flash("Invalid group or access denied.", 'danger')
            return redirect(url_for('dashboard'))
        
        progress = max(0, min(100, progress))
        
        details = ProjectDetails.query.filter_by(group_id=group.id).first()
        if not details:
            details = ProjectDetails(group_id=group.id, progress=progress)
            db.session.add(details)
        else:
            details.progress = progress
        
        db.session.commit()
        flash(f"Progress for {group.group_id} updated to {progress}%.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/supervisor/evaluate_group', methods=['POST'])
@login_required
def supervisor_evaluate_group():
    if current_user.role != 'supervisor':
        flash("Access denied.", 'danger')
        return redirect(url_for('dashboard'))
    
    group_id = request.form.get('group_id')
    status = request.form.get('status', 'Pending')
    feedback = request.form.get('feedback', '').strip()
    student_feedback = request.form.get('student_feedback', '').strip()
    
    try:
        group = StudentGroup.query.get(group_id)
        if not group or group.supervisor_id != current_user.id:
            flash("Invalid group or access denied.", 'danger')
            return redirect(url_for('dashboard'))
        
        if status not in ('Pending', 'Accepted', 'Conditionally Accepted', 'Deferred'):
            flash("Invalid evaluation status.", 'danger')
            return redirect(url_for('dashboard'))
        
        project_status = ProjectStatus.query.filter_by(group_id=group.id, teacher_id=current_user.id).first()
        if not project_status:
            project_status = ProjectStatus(
                group_id=group.id,
                teacher_id=current_user.id,
                status=status,
                feedback=feedback,
                student_feedback=student_feedback
            )
            db.session.add(project_status)
        else:
            project_status.status = status
            project_status.feedback = feedback
            project_status.student_feedback = student_feedback
        
        db.session.commit()
        logger.info(f"Group {group.group_id} evaluated as '{status}' by {current_user.email}")
        flash(f"Group {group.group_id} evaluated as '{status}'.", 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Error: {str(e)}", 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/seed_data')
def seed_data():
    # Only run this in development
    if app.config['ENV'] != 'production':
        # Create sample student groups for the current user
        if current_user.is_authenticated and current_user.role == 'teacher':
            # Check if groups already exist
            existing_groups = StudentGroup.query.filter_by(supervisor_id=current_user.id).all()
            if not existing_groups:
                groups = [
                    StudentGroup(group_id="G23", project_title="Smart Healthcare Monitoring System", supervisor_id=current_user.id),
                    StudentGroup(group_id="G17", project_title="Cloud-based Inventory System", supervisor_id=current_user.id),
                    StudentGroup(group_id="G19", project_title="Machine Learning for Financial Analysis", supervisor_id=current_user.id),
                    StudentGroup(group_id="G21", project_title="E-commerce Recommendation Engine", supervisor_id=current_user.id),
                    StudentGroup(group_id="G24", project_title="Blockchain for Supply Chain Tracking", supervisor_id=current_user.id),
                    StudentGroup(group_id="G27", project_title="Social Media Sentiment Analysis", supervisor_id=current_user.id),
                    StudentGroup(group_id="G29", project_title="Virtual Reality Campus Tour", supervisor_id=current_user.id)
                ]
                db.session.add_all(groups)
                db.session.commit()
                
                # Add sample remarks
                group1 = StudentGroup.query.filter_by(group_id="G23").first()
                group2 = StudentGroup.query.filter_by(group_id="G17").first()
                
                remarks = [
                    Remark(content="The team is making excellent progress with their data collection module.", 
                           teacher_id=current_user.id, group_id=group1.id),
                    Remark(content="Need to improve API documentation and integration testing.", 
                           teacher_id=current_user.id, group_id=group2.id),
                    Remark(content="Weekly progress reports are well-structured. Keep up the good work!", 
                           teacher_id=current_user.id, group_id=group1.id)
                ]
                db.session.add_all(remarks)
                db.session.commit()
                
                flash("Sample data created successfully!")
            else:
                flash("Sample data already exists!")
        else:
            flash("You need to be logged in as a teacher to create sample data.")
        return redirect(url_for('dashboard'))
    flash("This operation is not permitted in production.")
    return redirect(url_for('index'))

@app.route('/login/google')
def login_google():
    if not google or not OAUTH_CONFIGURED:
        flash('Google OAuth is not configured. Please set GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET in your .env file.', 'warning')
        return redirect(url_for('login'))
    
    # Require role selection before Google login
    selected_role = request.args.get('role', '').strip()
    if not selected_role or selected_role not in ('student', 'faculty', 'teacher', 'supervisor', 'admin'):
        flash('Please select your role before signing in with Google.', 'warning')
        return redirect(url_for('login'))
    
    # Store selected role in session for use after OAuth callback
    session['google_login_role'] = selected_role
    
    try:
        # Use localhost instead of 127.0.0.1 for Google OAuth compatibility
        redirect_uri = url_for('authorize', _external=True).replace('127.0.0.1', 'localhost')
        return google.authorize_redirect(redirect_uri)
    except Exception as e:
        flash(f'Google authentication error: {str(e)}. Please try again or use regular login.', 'danger')
        return redirect(url_for('login'))

@app.route('/authorize')
def authorize():
    if not google or not OAUTH_CONFIGURED:
        flash('Google OAuth is not properly configured.', 'warning')
        return redirect(url_for('login'))
    
    try:
        token = google.authorize_access_token()
        user_info = token.get('userinfo')
        
        if not user_info:
            flash('Failed to retrieve user information from Google. Please try again.', 'danger')
            return redirect(url_for('login'))
        
        # Get the role that was selected before Google login
        selected_role = session.pop('google_login_role', None)
        if not selected_role:
            flash('Role selection is required. Please select your role and try again.', 'warning')
            return redirect(url_for('login'))
        
        # Check if user exists with the given Google ID
        user = User.query.filter_by(google_id=user_info['sub']).first()
        
        if not user:
            # Check if a user exists with the same email
            user = User.query.filter_by(email=user_info['email']).first()
            
            if user:
                # Verify the selected role matches the existing user's role
                user_role_normalized = 'faculty' if user.role == 'teacher' else user.role
                selected_role_normalized = 'faculty' if selected_role == 'teacher' else selected_role
                
                if user_role_normalized != selected_role_normalized:
                    flash(f'Invalid role selected. You are registered as a {user.role}.', 'danger')
                    return redirect(url_for('login'))
                
                # Associate Google ID with the existing account
                user.google_id = user_info['sub']
                db.session.commit()
            else:
                # Create a new user with the selected role
                user = User(
                    email=user_info['email'],
                    first_name=user_info.get('given_name', ''),
                    last_name=user_info.get('family_name', ''),
                    role=selected_role,
                    google_id=user_info['sub']
                )
                db.session.add(user)
                db.session.commit()
        else:
            # Existing Google-linked user — verify role matches
            user_role_normalized = 'faculty' if user.role == 'teacher' else user.role
            selected_role_normalized = 'faculty' if selected_role == 'teacher' else selected_role
            
            if user_role_normalized != selected_role_normalized:
                flash(f'Invalid role selected. You are registered as a {user.role}.', 'danger')
                return redirect(url_for('login'))
        
        # Log the user in
        login_user(user)
        return redirect(url_for('dashboard'))
    except Exception as e:
        flash(f'Google login failed: {str(e)}. Please use regular login.', 'danger')
        return redirect(url_for('login'))

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email')
        user = User.query.filter_by(email=email).first()
        
        if user:
            # Generate a token
            token = user.generate_reset_token()
            
            # Create the reset link
            reset_link = url_for('reset_password', token=token, email=email, _external=True)
            
            # Create the email message
            subject = "Password Reset Request"
            body = f"""
To reset your password, visit the following link:
{reset_link}

If you did not make this request, simply ignore this email and no changes will be made.
"""
            # Send the email using our reliable function
            success, message = send_email(email, subject, body)
            
            if success:
                # Show a detailed message regardless of whether we're in debug mode
                # This makes it clearer for the user what they need to do
                if "saved to file" in message:
                    file_path = message.split(": ")[1]
                    flash(f'Due to email restrictions, your reset link has been saved locally instead of being sent. Please check this file for your reset link: {file_path}', 'warning')
                    flash(f'Reset link: <a href="{reset_link}">{reset_link}</a>', 'info')
                    return render_template('forgot_password.html')
                else:
                    flash(f'A password reset link has been sent to your email {email}. Please check your inbox (and spam folder).', 'success')
                    # Still provide the direct link in development mode
                    if app.debug:
                        flash(f'For testing purposes, you can also use this direct link: <a href="{reset_link}">{reset_link}</a>', 'info')
                        return render_template('forgot_password.html')
                    return redirect(url_for('login'))
            else:
                # Show error message if email sending failed
                flash(f'Error sending email: {message}', 'danger')
                # Always provide the reset link in development mode for testing
                if app.debug:
                    flash(f'Reset link: <a href="{reset_link}">{reset_link}</a>', 'info')
                
                return render_template('forgot_password.html')
        else:
            flash('No account found with that email address.', 'danger')
    
    return render_template('forgot_password.html')

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    email = request.args.get('email')
    user = User.query.filter_by(email=email).first()
    
    if not user or not user.verify_reset_token(token):
        flash('Invalid or expired reset token.', 'danger')
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        
        if password != confirm_password:
            flash('Passwords do not match.', 'danger')
            return render_template('reset_password.html', token=token, email=email)
        
        user.set_password(password)
        user.clear_reset_token()
        db.session.commit()
        
        flash('Your password has been updated! You can now log in with your new password.', 'success')
        return redirect(url_for('login'))
    
    return render_template('reset_password.html', token=token, email=email)

@app.route('/email-instructions')
def email_instructions():
    """Show instructions on how to find and use password reset emails in development mode."""
    return render_template('email_instructions.html')

# Admin User Management
@app.route('/admin/users')
@login_required
def admin_users():
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    # User management is in the admin dashboard (dashboard_admin_modern) #users section
    return redirect(url_for('dashboard_admin') + '#users')

@app.route('/admin/add_user', methods=['POST'])
@login_required
def admin_add_user():
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    email = request.form.get('email')
    first_name = request.form.get('firstName')
    last_name = request.form.get('lastName')
    role = request.form.get('role')
    password = request.form.get('password')
    
    # Check if user already exists
    if User.query.filter_by(email=email).first():
        flash('Email already registered', 'danger')
        return redirect(url_for('dashboard') + '#users')
    
    # Create new user
    user = User(
        email=email,
        first_name=first_name,
        last_name=last_name,
        role=role
    )
    user.set_password(password)
    
    db.session.add(user)
    db.session.commit()
    
    flash(f'User {first_name} {last_name} added successfully', 'success')
    return redirect(url_for('dashboard') + '#users')

@app.route('/admin/edit_user/<int:user_id>', methods=['POST'])
@login_required
def admin_edit_user(user_id):
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    user = User.query.get_or_404(user_id)
    
    try:
        new_email = request.form.get('email')
        # Check for duplicate email (excluding current user)
        if new_email != user.email:
            existing = User.query.filter_by(email=new_email).first()
            if existing:
                flash(f'Email {new_email} is already in use by another user', 'danger')
                return redirect(url_for('dashboard') + '#users')
        
        user.email = new_email
        user.first_name = request.form.get('firstName')
        user.last_name = request.form.get('lastName')
        user.role = request.form.get('role')
        
        # Only update password if provided
        if request.form.get('password'):
            user.set_password(request.form.get('password'))
        
        db.session.commit()
        flash(f'User {user.first_name} {user.last_name} updated successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating user: {str(e)}', 'danger')
    return redirect(url_for('dashboard') + '#users')

@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
def admin_delete_user(user_id):
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    user = User.query.get_or_404(user_id)
    
    try:
        # Check if this user has any group memberships
        # Define inspector to check if table exists
        inspector = inspect(db.engine)
        
        # Only attempt to delete group memberships if the table exists
        if 'group_member' in inspector.get_table_names():
            # Delete any group memberships
            GroupMember.query.filter_by(user_id=user.id).delete()
        
        # Delete user remarks if they were faculty/teacher
        if user.role in ('faculty', 'teacher'):
            Remark.query.filter_by(teacher_id=user.id).delete()
        
        # Delete ProjectStatus records where this user is the teacher/evaluator
        ProjectStatus.query.filter_by(teacher_id=user.id).delete()
        
        # Handle groups if they were a supervisor
        if user.role == 'supervisor':
            # Just set supervisor_id to NULL for any groups they supervised
            groups = StudentGroup.query.filter_by(supervisor_id=user.id).all()
            for group in groups:
                group.supervisor_id = None
        
        # Now delete the user
        db.session.delete(user)
        db.session.commit()
        
        flash(f'User {user.email} has been deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting user: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard') + '#users')

# Admin Project Management
@app.route('/admin/add_project', methods=['POST'])
@login_required
def admin_add_project():
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    try:
        project_title = request.form.get('project_title', '').strip()
        group_id = request.form.get('group_id', '').strip()
        supervisor_id = request.form.get('supervisor_id')
        student_ids = request.form.getlist('student_ids')
        
        # Validate project title is provided (required)
        if not project_title:
            flash('Project Title is required', 'danger')
            return redirect(url_for('dashboard'))
        
        # Auto-generate Group ID if not provided
        if not group_id:
            # Generate group ID from project title and timestamp
            import time
            group_id = f"G{int(time.time()) % 10000}"
        
        # Check if group_id already exists
        if StudentGroup.query.filter_by(group_id=group_id).first():
            flash(f'Group ID {group_id} already exists. Please use a different ID.', 'danger')
            return redirect(url_for('dashboard'))
        
        # Create new project group
        group = StudentGroup(
            group_id=group_id,
            project_title=project_title,
            supervisor_id=int(supervisor_id) if supervisor_id else None
        )
        
        db.session.add(group)
        db.session.flush()  # Flush to get the group ID without committing
        
        # Add selected students to the group (optional)
        if student_ids:
            for student_id in student_ids:
                group_member = GroupMember(user_id=int(student_id), group_id=group.id)
                db.session.add(group_member)
        
        db.session.commit()
        
        # Build success message
        message = f'Project "{project_title}" (ID: {group_id}) created successfully'
        
        if student_ids:
            message += f' with {len(student_ids)} student(s)'
        
        if supervisor_id:
            supervisor = User.query.get(int(supervisor_id))
            if supervisor:
                message += f' - Supervisor: {supervisor.first_name} {supervisor.last_name}'
        
        flash(message, 'success')
        
    except Exception as e:
        flash(f'Error creating project: {str(e)}', 'danger')
        db.session.rollback()
    
    # Always redirect back to admin dashboard - regardless of success or failure
    return redirect(url_for('dashboard'))

@app.route('/admin/edit_project/<int:project_id>', methods=['POST'])
@login_required
def admin_edit_project(project_id):
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    group = StudentGroup.query.get_or_404(project_id)
    
    try:
        group.group_id = request.form.get('group_id')
        group.project_title = request.form.get('project_title')
        supervisor_id = request.form.get('supervisor_id')
        group.supervisor_id = int(supervisor_id) if supervisor_id else None
        
        db.session.commit()
        flash(f'Project "{group.project_title}" updated successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating project: {str(e)}', 'danger')
    return redirect(url_for('dashboard'))

@app.route('/admin/delete_project/<int:project_id>', methods=['POST'])
@login_required
def admin_delete_project(project_id):
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    group = StudentGroup.query.get_or_404(project_id)
    project_title = group.project_title
    
    try:
        # Delete associated group members
        GroupMember.query.filter_by(group_id=group.id).delete()
        
        # Delete associated remarks
        Remark.query.filter_by(group_id=group.id).delete()
        
        # Delete associated project statuses
        ProjectStatus.query.filter_by(group_id=group.id).delete()
        
        # Delete the group project itself
        db.session.delete(group)
        db.session.commit()
        
        flash(f'Group project "{project_title}" has been deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting group project: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard'))

# Admin Group Management
@app.route('/admin/assign_member', methods=['POST'])
@login_required
def admin_assign_member():
    if current_user.role != "admin":
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Access denied'})
        flash("Access denied.")
        return redirect(url_for('index'))
    
    group_id = request.form.get('group_id')
    student_ids = request.form.getlist('student_ids')  # Get list of multiple students
    
    # Validate inputs
    if not group_id or not student_ids:
        message = 'Missing required information'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': message})
        flash(message, 'danger')
        return redirect(url_for('dashboard'))
    
    # Find the group
    group = StudentGroup.query.filter_by(id=group_id).first()
    if not group:
        message = 'Group not found'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': message})
        flash(message, 'danger')
        return redirect(url_for('dashboard'))
    
    # Process multiple students
    added_students = []
    already_in_group = []
    not_found = []
    
    for student_id in student_ids:
        # Find the student
        student = User.query.filter_by(id=student_id, role='student').first()
        if not student:
            not_found.append(student_id)
            continue
        
        # Check if already in group
        if GroupMember.query.filter_by(user_id=student.id, group_id=group.id).first():
            already_in_group.append(f"{student.first_name} {student.last_name}")
            continue
        
        # Add student to group
        group_member = GroupMember(user_id=student.id, group_id=group.id)
        db.session.add(group_member)
        added_students.append(f"{student.first_name} {student.last_name}")
    
    # Commit all changes
    if added_students or already_in_group:
        db.session.commit()
    
    # Build response message
    message_parts = []
    if added_students:
        message_parts.append(f"✓ Added {len(added_students)} student(s): {', '.join(added_students)}")
    if already_in_group:
        message_parts.append(f"⚠️ {len(already_in_group)} student(s) already in group: {', '.join(already_in_group)}")
    if not_found:
        message_parts.append(f"❌ {len(not_found)} student(s) not found")
    
    full_message = " | ".join(message_parts) if message_parts else "No changes made"
    message_type = 'success' if added_students else ('warning' if already_in_group else 'danger')
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': bool(added_students), 'message': full_message})
    
    flash(full_message, message_type)
    return redirect(url_for('dashboard'))

# Get the list of students in a group
@app.route('/admin/group_members/<int:group_id>')
@login_required
def admin_group_members(group_id):
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    group = StudentGroup.query.get_or_404(group_id)
    
    # Get all members via relationship
    members = User.query.join(GroupMember).filter(GroupMember.group_id == group.id).all()
    
    return jsonify([{
        'id': member.id,
        'name': f"{member.first_name} {member.last_name}",
        'email': member.email
    } for member in members])

# Remove a student from a group
@app.route('/admin/remove_member', methods=['POST'])
@login_required
def admin_remove_member():
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    group_id = request.form.get('group_id')
    student_id = request.form.get('student_id')
    
    # Validate inputs
    if not group_id or not student_id:
        flash('Missing required information', 'danger')
        return redirect(url_for('dashboard'))
    
    # Delete the membership
    membership = GroupMember.query.filter_by(user_id=student_id, group_id=group_id).first()
    if membership:
        db.session.delete(membership)
        db.session.commit()
        flash('Student removed from group', 'success')
    else:
        flash('Student not found in group', 'danger')
    
    return redirect(url_for('dashboard'))

# Admin System Settings
@app.route('/admin/save_settings', methods=['POST'])
@login_required
def admin_save_settings():
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    # Process general settings
    system_name = request.form.get('systemName')
    timezone = request.form.get('timezone')
    date_format = request.form.get('dateFormat')
    
    # In a real app, you would save these settings to the database or a config file
    # For now, we'll just show a success message
    flash('System settings updated successfully', 'success')
    return redirect(url_for('dashboard'))

# Admin Teacher Username Management
@app.route('/admin/teacher_usernames')
@login_required
def admin_teacher_usernames():
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    usernames = TeacherUsername.query.all()
    return jsonify([{
        'id': username.id,
        'username': username.username,
        'is_used': username.is_used,
        'user_id': username.user_id,
        'user_name': f"{username.user.first_name} {username.user.last_name}" if username.user else None,
        'created_at': username.created_at.isoformat() if username.created_at else None
    } for username in usernames])

@app.route('/admin/add_teacher_username', methods=['POST'])
@login_required
def admin_add_teacher_username():
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    raw_username = request.form.get('username')
    
    # Validate username
    if not raw_username or not raw_username.strip():
        flash('Username is required', 'danger')
        return redirect(url_for('dashboard'))
    
    username = raw_username.strip().lower()
    
    if not username:
        flash('Username is required', 'danger')
        return redirect(url_for('dashboard'))
    
    # Check if username already exists
    if TeacherUsername.query.filter_by(username=username).first():
        flash(f'Username {username} already exists', 'danger')
        return redirect(url_for('dashboard'))
    
    # Create new teacher username
    teacher_username = TeacherUsername(username=username)
    db.session.add(teacher_username)
    db.session.commit()
    
    flash(f'Teacher username {username} added successfully', 'success')
    return redirect(url_for('dashboard'))

@app.route('/admin/delete_teacher_username/<int:username_id>', methods=['POST'])
@login_required
def admin_delete_teacher_username(username_id):
    if current_user.role != "admin":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    username = TeacherUsername.query.get_or_404(username_id)
    
    # Check if username is already in use
    if username.is_used:
        flash('Cannot delete username that is already in use', 'danger')
        return redirect(url_for('dashboard'))
    
    db.session.delete(username)
    db.session.commit()
    
    flash(f'Teacher username {username.username} has been deleted', 'success')
    return redirect(url_for('dashboard'))

# Scheduling Routes
@app.route("/admin/scheduling", methods=["GET"])
@login_required
def admin_scheduling():
    if current_user.role != 'admin':
        flash('You do not have permission to access this page', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get all rooms, time slots, faculty, and groups
    rooms = Room.query.all()
    time_slots = TimeSlot.query.order_by(TimeSlot.day, TimeSlot.start_time).all()
    faculty = User.query.filter(User.role.in_(['teacher', 'faculty'])).all()
    groups = StudentGroup.query.all()
    
    # Get existing schedules
    teacher_schedules = TeacherSchedule.query.all()
    room_schedules = RoomSchedule.query.all()
    
    # Organize time slots by day for easier rendering
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    time_slots_by_day = {day: [] for day in days}
    
    for slot in time_slots:
        time_slots_by_day[slot.day].append(slot)
    
    # Create data structures for easier rendering in the template
    teacher_schedule_map = {}
    for ts in teacher_schedules:
        key = f"{ts.teacher_id}_{ts.time_slot_id}"
        teacher_schedule_map[key] = ts
    
    room_schedule_map = {}
    for rs in room_schedules:
        key = f"{rs.room_id}_{rs.time_slot_id}"
        room_schedule_map[key] = rs
    
    return render_template(
        'schedule_admin.html',
        rooms=rooms,
        time_slots=time_slots,
        time_slots_by_day=time_slots_by_day,
        faculty=faculty,
        groups=groups,
        teacher_schedule_map=teacher_schedule_map,
        room_schedule_map=room_schedule_map,
        days=days
    )

@app.route("/admin/add_teacher_schedule", methods=["POST"])
@login_required
def admin_add_teacher_schedule():
    if current_user.role != 'admin':
        flash('You do not have permission to manage schedules', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        teacher_id = request.form.get('teacher_id')
        time_slot_id = request.form.get('time_slot_id')
        subject = request.form.get('subject')
        room_name = request.form.get('room_name')  # Get the room name
        class_name = request.form.get('class_name')  # Get the class name
        group_id = request.form.get('group_id')  # Get the group ID
        
        if not teacher_id or not time_slot_id or not subject or not room_name or not class_name or not group_id:
            flash('All fields are required', 'danger')
            return redirect(url_for('admin_scheduling'))
        
        # Verify teacher exists
        teacher = User.query.get(teacher_id)
        if not teacher:
            flash(f'Teacher with ID {teacher_id} does not exist', 'danger')
            return redirect(url_for('admin_scheduling'))
        
        # Verify time slot exists
        time_slot = TimeSlot.query.get(time_slot_id)
        if not time_slot:
            flash(f'Time slot with ID {time_slot_id} does not exist', 'danger')
            return redirect(url_for('admin_scheduling'))
        
        # Verify group exists
        group = StudentGroup.query.get(group_id)
        if not group:
            flash(f'Group with ID {group_id} does not exist', 'danger')
            return redirect(url_for('admin_scheduling'))
        
        # Find or create room with the given name
        room = Room.query.filter_by(name=room_name).first()
        if not room:
            room = Room(name=room_name, capacity=30)  # Default capacity of 30
            db.session.add(room)
            db.session.commit()
        
        room_id = room.id
        
        # Check if the teacher already has a schedule for this time slot
        existing_teacher_schedule = TeacherSchedule.query.filter_by(
            teacher_id=teacher_id,
            time_slot_id=time_slot_id
        ).first()
        
        if existing_teacher_schedule:
            flash('Teacher already has a schedule for this time slot', 'danger')
            return redirect(url_for('admin_scheduling'))
        
        # Check if the room already has a schedule for this time slot
        existing_room_schedule = RoomSchedule.query.filter_by(
            room_id=room_id,
            time_slot_id=time_slot_id
        ).first()
        
        if existing_room_schedule:
            flash('Room already has a schedule for this time slot', 'danger')
            return redirect(url_for('admin_scheduling'))
        
        # Create new teacher schedule
        teacher_schedule = TeacherSchedule(
            teacher_id=teacher_id,
            time_slot_id=time_slot_id,
            group_id=group_id,
            subject=subject,
            class_name=class_name
        )
        
        # Create new room schedule
        room_schedule = RoomSchedule(
            room_id=room_id,
            time_slot_id=time_slot_id,
            group_id=group_id,
            class_name=class_name
        )
        
        db.session.add(teacher_schedule)
        db.session.add(room_schedule)
        db.session.commit()
        
        logger.info(f"Schedule added by {current_user.email}: teacher={teacher_id}, slot={time_slot_id}, group={group_id}")
        flash('Teacher and room schedule added successfully', 'success')
        return redirect(url_for('admin_scheduling'))
    
    except ValueError as ve:
        db.session.rollback()
        logger.error(f"Validation error adding schedule: {str(ve)}")
        flash(f"Error adding schedule: {str(ve)}", 'danger')
        return redirect(url_for('admin_scheduling'))
    except Exception as e:
        db.session.rollback()
        logger.error(f"Unexpected error adding schedule: {str(e)}")
        flash(f"Unexpected error: {str(e)}", 'danger')
        return redirect(url_for('admin_scheduling'))

@app.route("/admin/delete_teacher_schedule/<int:schedule_id>", methods=["POST"])
@login_required
def admin_delete_teacher_schedule(schedule_id):
    if current_user.role != 'admin':
        flash('You do not have permission to manage schedules', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        # Get the teacher schedule
        schedule = TeacherSchedule.query.get_or_404(schedule_id)
        
        # Find any corresponding room schedules with the same time slot
        room_schedules = RoomSchedule.query.filter_by(time_slot_id=schedule.time_slot_id).all()
        
        # Delete the teacher schedule
        db.session.delete(schedule)
        
        # Delete the corresponding room schedules
        for room_schedule in room_schedules:
            db.session.delete(room_schedule)
        
        db.session.commit()
        
        flash('Teacher schedule and related room schedules deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting teacher schedule: {str(e)}")
        flash(f'Error deleting schedule: {str(e)}', 'danger')
        
    return redirect(url_for('admin_scheduling'))

@app.route("/admin/add_room_schedule", methods=["POST"])
@login_required
def admin_add_room_schedule():
    if current_user.role != 'admin':
        flash('You do not have permission to manage schedules', 'danger')
        return redirect(url_for('dashboard'))
    
    room_id = request.form.get('room_id')
    time_slot_id = request.form.get('time_slot_id')
    class_name = request.form.get('class_name')
    teacher_id = request.form.get('teacher_id')  # Get teacher_id if provided
    subject = request.form.get('subject')  # Get subject if provided
    
    if not room_id or not time_slot_id or not class_name:
        flash('Room, time slot, and class name are required', 'danger')
        return redirect(url_for('admin_scheduling'))
    
    # Check if the room already has a schedule for this time slot
    existing_room_schedule = RoomSchedule.query.filter_by(
        room_id=room_id,
        time_slot_id=time_slot_id
    ).first()
    
    if existing_room_schedule:
        flash('Room already has a schedule for this time slot', 'danger')
        return redirect(url_for('admin_scheduling'))
    
    # Create new room schedule
    room_schedule = RoomSchedule(
        room_id=room_id,
        time_slot_id=time_slot_id,
        class_name=class_name
    )
    
    db.session.add(room_schedule)
    
    # If teacher and subject are provided, create teacher schedule too
    if teacher_id and subject:
        # Check if the teacher already has a schedule for this time slot
        existing_teacher_schedule = TeacherSchedule.query.filter_by(
            teacher_id=teacher_id,
            time_slot_id=time_slot_id
        ).first()
        
        if not existing_teacher_schedule:
            teacher_schedule = TeacherSchedule(
                teacher_id=teacher_id,
                time_slot_id=time_slot_id,
                subject=subject
            )
            db.session.add(teacher_schedule)
            db.session.commit()
            flash('Room schedule and teacher schedule added successfully', 'success')
        else:
            db.session.commit()
            flash('Room schedule added, but teacher already has a schedule for this time slot', 'warning')
    else:
        db.session.commit()
        flash('Room schedule added successfully', 'success')
    
    return redirect(url_for('admin_scheduling'))

@app.route("/admin/delete_room_schedule/<int:schedule_id>", methods=["POST"])
@login_required
def admin_delete_room_schedule(schedule_id):
    if current_user.role != 'admin':
        flash('You do not have permission to manage schedules', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        # Get the room schedule
        schedule = RoomSchedule.query.get_or_404(schedule_id)
        
        # Delete the room schedule
        db.session.delete(schedule)
        db.session.commit()
        
        flash('Room schedule deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting room schedule: {str(e)}")
        flash(f'Error deleting room schedule: {str(e)}', 'danger')
        
    return redirect(url_for('admin_scheduling'))

# Room Management Routes
@app.route("/admin/add_room", methods=["POST"])
@login_required
def admin_add_room():
    if current_user.role != 'admin':
        flash('You do not have permission to manage rooms', 'danger')
        return redirect(url_for('dashboard'))
    
    room_name = request.form.get('room_name', '').strip()
    capacity = request.form.get('capacity', 30, type=int)
    description = request.form.get('description', '').strip()
    
    if not room_name:
        flash('Room name is required', 'danger')
        return redirect(url_for('admin_scheduling'))
    
    # Check if room already exists
    existing_room = Room.query.filter_by(name=room_name).first()
    if existing_room:
        flash(f'Room "{room_name}" already exists', 'warning')
        return redirect(url_for('admin_scheduling'))
    
    try:
        room = Room(name=room_name, capacity=capacity, description=description if description else None)
        db.session.add(room)
        db.session.commit()
        flash(f'Room "{room_name}" added successfully', 'success')
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error adding room: {str(e)}")
        flash(f'Error adding room: {str(e)}', 'danger')
    
    return redirect(url_for('admin_scheduling'))


@app.route("/admin/delete_room/<int:room_id>", methods=["POST"])
@login_required
def admin_delete_room(room_id):
    if current_user.role != 'admin':
        flash('You do not have permission to manage rooms', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        room = Room.query.get_or_404(room_id)
        room_name = room.name
        
        # Delete all room schedules for this room first
        RoomSchedule.query.filter_by(room_id=room_id).delete()
        
        db.session.delete(room)
        db.session.commit()
        flash(f'Room "{room_name}" and its schedules deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting room: {str(e)}")
        flash(f'Error deleting room: {str(e)}', 'danger')
    
    return redirect(url_for('admin_scheduling'))


# Viva Scheduling Routes
@app.route("/admin/viva_scheduling", methods=["GET"])
@login_required
def admin_viva_scheduling():
    if current_user.role != 'admin':
        flash('You do not have permission to access this page', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get all student groups, teachers, and rooms
    groups = StudentGroup.query.all()
    teachers = User.query.filter(User.role.in_(['teacher', 'faculty', 'supervisor'])).all()
    rooms = Room.query.all()
    
    # Get all scheduled vivas
    vivas = Viva.query.order_by(Viva.scheduled_date, Viva.scheduled_time).all()
    
    # Get today's date for minimum date in date picker
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    
    return render_template(
        'viva_scheduling.html',
        groups=groups,
        teachers=teachers,
        rooms=rooms,
        vivas=vivas,
        today=today
    )

@app.route("/admin/schedule_viva", methods=["POST"])
@login_required
def admin_schedule_viva():
    if current_user.role != 'admin':
        flash('You do not have permission to manage vivas', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get form data
    group_id = request.form.get('group_id')
    teacher_id = request.form.get('teacher_id')
    scheduled_date_str = request.form.get('scheduled_date')
    scheduled_time_str = request.form.get('scheduled_time')
    duration_minutes = request.form.get('duration_minutes', 30)
    location = request.form.get('location')
    notes = request.form.get('notes', '')
    
    # Validate required fields
    if not group_id or not teacher_id or not scheduled_date_str or not scheduled_time_str or not location:
        flash('All fields are required except notes', 'danger')
        return redirect(url_for('admin_viva_scheduling'))
    
    # Parse the date
    scheduled_date = datetime.datetime.strptime(scheduled_date_str, '%Y-%m-%d').date()
    
    # Create new viva
    viva = Viva(
        group_id=group_id,
        teacher_id=teacher_id,
        scheduled_date=scheduled_date,
        scheduled_time=scheduled_time_str,
        duration_minutes=duration_minutes,
        location=location,
        notes=notes,
        status='Scheduled'
    )
    
    db.session.add(viva)
    db.session.commit()
    
    # Also create/update ProjectStatus to show in faculty dashboard
    project_status = ProjectStatus.query.filter_by(group_id=group_id, teacher_id=teacher_id).first()
    if not project_status:
        project_status = ProjectStatus(
            group_id=group_id,
            teacher_id=teacher_id,
            status='Pending'
        )
        db.session.add(project_status)
        db.session.commit()
    
    flash('Viva scheduled successfully', 'success')
    return redirect(url_for('admin_viva_scheduling'))

@app.route("/admin/delete_viva/<int:viva_id>", methods=["POST"])
@login_required
def admin_delete_viva(viva_id):
    if current_user.role != 'admin':
        flash('You do not have permission to manage vivas', 'danger')
        return redirect(url_for('dashboard'))
    
    viva = Viva.query.get_or_404(viva_id)
    db.session.delete(viva)
    db.session.commit()
    
    flash('Viva deleted successfully', 'success')
    return redirect(url_for('admin_viva_scheduling'))

@app.route("/api/check_availability", methods=["GET"])
@login_required
def check_availability():
    """API to check teacher and room availability for scheduling vivas."""
    # Get request parameters
    teacher_id = request.args.get('teacher_id')
    date_str = request.args.get('date')
    room_id = None
    room_name = request.args.get('room_id')
    
    if not teacher_id or not date_str:
        return jsonify({
            'error': 'Missing required parameters',
            'available_slots': []
        }), 400
    
    try:
        # Parse the date string
        check_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        day_of_week = check_date.strftime('%A')  # Convert date to day name (Monday, Tuesday, etc.)
        
        # Get all time slots
        time_slots = TimeSlot.query.filter_by(day=day_of_week).order_by(TimeSlot.start_time).all()
        time_slot_map = {ts.id: ts for ts in time_slots}
        
        # Get teacher's regular classes for that day of the week
        teacher_classes = TeacherSchedule.query.filter_by(teacher_id=teacher_id).all()
        teacher_busy_times = []
        
        # Add teacher's classes to busy times
        for cls in teacher_classes:
            time_slot = time_slot_map.get(cls.time_slot_id)
            if time_slot and time_slot.day == day_of_week:
                teacher_busy_times.append({
                    'start': time_slot.start_time,
                    'end': time_slot.end_time,
                    'reason': f"Teaching {cls.subject}"
                })
        
        # Get teacher's existing vivas on that date
        teacher_vivas = Viva.query.filter_by(
            teacher_id=teacher_id,
            scheduled_date=check_date,
            status='Scheduled'
        ).all()
        
        # Add teacher's vivas to busy times
        for viva in teacher_vivas:
            start_time = viva.scheduled_time
            # Calculate end time based on duration
            start_dt = datetime.datetime.strptime(start_time, '%H:%M')
            end_dt = start_dt + datetime.timedelta(minutes=int(viva.duration_minutes))
            end_time = end_dt.strftime('%H:%M')
            
            teacher_busy_times.append({
                'start': start_time,
                'end': end_time,
                'reason': f"Evaluating group {viva.group.group_id}"
            })
        
        # Check room availability if room_name is provided
        room_busy_times = []
        if room_name:
            # Get room's ID by name
            room = Room.query.filter_by(name=room_name).first()
            if room:
                room_id = room.id
            
            if room_id:
                # Get room's regular classes
                room_classes = RoomSchedule.query.filter_by(room_id=room_id).all()
                
                # Add room's classes to busy times
                for cls in room_classes:
                    time_slot = time_slot_map.get(cls.time_slot_id)
                    if time_slot and time_slot.day == day_of_week:
                        room_busy_times.append({
                            'start': time_slot.start_time,
                            'end': time_slot.end_time,
                            'reason': f"Room booked for {cls.class_name}"
                        })
                
                # Get room's existing vivas
                room_vivas = Viva.query.filter_by(
                    location=room.name,
                    scheduled_date=check_date,
                    status='Scheduled'
                ).all()
                
                # Add room's vivas to busy times
                for viva in room_vivas:
                    start_time = viva.scheduled_time
                    # Calculate end time based on duration
                    start_dt = datetime.datetime.strptime(start_time, '%H:%M')
                    end_dt = start_dt + datetime.timedelta(minutes=int(viva.duration_minutes))
                    end_time = end_dt.strftime('%H:%M')
                    
                    room_busy_times.append({
                        'start': start_time,
                        'end': end_time,
                        'reason': f"Room booked for viva (Group {viva.group.group_id})"
                    })
        
        # Combine all busy times
        busy_times = teacher_busy_times + room_busy_times
        
        # Generate available time slots (9 AM to 5 PM in 30-minute increments)
        all_slots = []
        current_time = datetime.datetime.strptime('09:00', '%H:%M')
        end_of_day = datetime.datetime.strptime('17:00', '%H:%M')
        
        while current_time < end_of_day:
            slot_start = current_time.strftime('%H:%M')
            current_time += datetime.timedelta(minutes=30)
            slot_end = current_time.strftime('%H:%M')
            
            # Check if this slot overlaps with any busy time
            is_available = True
            conflicts = []
            
            for busy in busy_times:
                busy_start = datetime.datetime.strptime(busy['start'], '%H:%M')
                busy_end = datetime.datetime.strptime(busy['end'], '%H:%M')
                
                # Fix the overlap detection logic - check if the slot overlaps with busy time
                slot_start_dt = datetime.datetime.strptime(slot_start, '%H:%M')
                slot_end_dt = datetime.datetime.strptime(slot_end, '%H:%M')
                
                # A slot overlaps if:
                # 1. Slot start is within busy period (busy_start <= slot_start < busy_end), OR
                # 2. Slot end is within busy period (busy_start < slot_end <= busy_end), OR
                # 3. Slot completely contains busy period (slot_start <= busy_start && slot_end >= busy_end)
                if ((busy_start <= slot_start_dt and slot_start_dt < busy_end) or
                    (busy_start < slot_end_dt and slot_end_dt <= busy_end) or
                    (slot_start_dt <= busy_start and slot_end_dt >= busy_end)):
                    is_available = False
                    conflicts.append(busy['reason'])
            
            all_slots.append({
                'start': slot_start,
                'end': slot_end,
                'available': is_available,
                'conflicts': conflicts
            })
        
        return jsonify({
            'busy_times': busy_times,
            'available_slots': [slot for slot in all_slots if slot['available']]
        })
    
    except Exception as e:
        return jsonify({
            'error': str(e),
            'available_slots': []
        }), 500

@app.route("/teacher/schedule", methods=["GET"])
@login_required
def teacher_schedule():
    if current_user.role != 'teacher':
        flash('You do not have permission to access this page', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get teacher's schedule
    schedules = TeacherSchedule.query.filter_by(teacher_id=current_user.id).all()
    time_slots = {ts.id: ts for ts in TimeSlot.query.all()}
    
    # Organize schedules by day for easier rendering
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    schedules_by_day = {day: [] for day in days}
    
    for schedule in schedules:
        time_slot = time_slots[schedule.time_slot_id]
        schedules_by_day[time_slot.day].append({
            'id': schedule.id,
            'time': f"{time_slot.start_time} - {time_slot.end_time}",
            'subject': schedule.subject
        })
    
    return render_template(
        'schedule_teacher.html',
        schedules_by_day=schedules_by_day,
        days=days
    )

@app.route("/student/schedule", methods=["GET"])
@login_required
def student_schedule():
    if current_user.role != 'student':
        flash('You do not have permission to access this page', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get all room schedules (in a real app, you'd filter by the student's enrolled classes)
    room_schedules = RoomSchedule.query.all()
    
    # Get all time slots and rooms for reference
    time_slots = {ts.id: ts for ts in TimeSlot.query.all()}
    rooms = {r.id: r for r in Room.query.all()}
    
    # Get teacher information for each schedule
    teacher_schedules = TeacherSchedule.query.all()
    teacher_map = {}
    for ts in teacher_schedules:
        key = f"{ts.time_slot_id}"
        teacher = User.query.get(ts.teacher_id)
        if teacher:
            teacher_map[key] = {
                'name': f"{teacher.first_name} {teacher.last_name}",
                'subject': ts.subject
            }
    
    # Organize schedules by day for easier rendering
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    schedules_by_day = {day: [] for day in days}
    
    for schedule in room_schedules:
        if schedule.time_slot_id in time_slots and schedule.room_id in rooms:
            time_slot = time_slots[schedule.time_slot_id]
            room = rooms[schedule.room_id]
            
            # Get teacher info if available
            teacher_info = teacher_map.get(f"{schedule.time_slot_id}", None)
            teacher_name = teacher_info['name'] if teacher_info else "No teacher assigned"
            subject = teacher_info['subject'] if teacher_info else schedule.class_name
            
            schedules_by_day[time_slot.day].append({
                'id': schedule.id,
                'time': f"{time_slot.start_time} - {time_slot.end_time}",
                'class_name': schedule.class_name,
                'room': room.name,
                'teacher': teacher_name,
                'subject': subject
            })
    
    # Sort each day's schedules by time
    for day in days:
        schedules_by_day[day].sort(key=lambda x: x['time'])
    
    return render_template(
        'schedule_student.html',
        schedules_by_day=schedules_by_day,
        days=days
    )

# Define a helper function to safely recreate database tables
def recreate_tables():
    print("Recreating database tables due to schema change...")
    db.drop_all()
    db.create_all()
    print("Database tables recreated.")
    
    try:
        # Check if admin user already exists
        admin = User.query.filter_by(email='admin@example.com').first()
        if not admin:
            # Create initial users
            admin = User(email='admin@example.com', first_name='Admin', last_name='User', role='admin')
            admin.set_password('admin123')
            db.session.add(admin)
            
            # Create a sample teacher
            teacher = User(email='teacher@example.com', first_name='John', last_name='Smith', role='faculty')
            teacher.set_password('teacher123')
            db.session.add(teacher)
            
            # Create a sample supervisor
            supervisor = User(email='supervisor@example.com', first_name='David', last_name='Johnson', role='supervisor')
            supervisor.set_password('supervisor123')
            db.session.add(supervisor)
            
            # Create a sample student
            student = User(email='student@example.com', first_name='Sarah', last_name='Johnson', role='student')
            student.set_password('student123')
            db.session.add(student)
            
            db.session.commit()
            print("Sample users created.")
    except Exception as e:
        db.session.rollback()
        print(f"Error creating sample users: {str(e)}")

# Helper function to fix the Viva table
def fix_viva_table():
    print("Trying to fix Viva table...")
    try:
        # Check if Viva table exists but has wrong schema
        inspector = inspect(db.engine)
        if 'viva' in inspector.get_table_names():
            # Get columns in the existing Viva table
            columns = [col['name'] for col in inspector.get_columns('viva')]
            
            # If scheduled_time is missing, recreate just the Viva table
            if 'scheduled_time' not in columns:
                print("Viva table exists but missing scheduled_time column. Recreating table...")
                Viva.__table__.drop(db.engine)
                Viva.__table__.create(db.engine)
                print("Viva table recreated successfully.")
            else:
                print("Viva table schema looks correct.")
        else:
            print("Viva table doesn't exist. Creating it...")
            Viva.__table__.create(db.engine)
            print("Viva table created successfully.")
        
        return True
    except Exception as e:
        print(f"Error fixing Viva table: {str(e)}")
        return False

@app.route('/teacher/assigned_groups')
@login_required
def teacher_assigned_groups():
    """View all project groups assigned to the teacher for evaluation"""
    # Ensure the user is a teacher
    if current_user.role != 'teacher':
        flash('Access denied. This page is only for teachers.', 'danger')
        return redirect(url_for('index'))
    
    # Get all status entries for this teacher
    status_entries = ProjectStatus.query.filter_by(teacher_id=current_user.id).all()
    
    # Get the unique groups
    group_ids = set(entry.group_id for entry in status_entries)
    groups = StudentGroup.query.filter(StudentGroup.id.in_(group_ids)).all()
    
    # Count statistics
    pending_count = sum(1 for entry in status_entries if entry.status == 'Pending')
    accepted_count = sum(1 for entry in status_entries if entry.status == 'Accepted')
    conditionally_accepted_count = sum(1 for entry in status_entries if entry.status == 'Conditionally Accepted')
    deferred_count = sum(1 for entry in status_entries if entry.status == 'Deferred')
    
    # Get upcoming viva examinations
    upcoming_vivas = Viva.query.filter_by(
        teacher_id=current_user.id,
        status='Scheduled'
    ).filter(Viva.scheduled_date >= datetime.datetime.now().date()).all()
    
    # Helper function to get status entry for a specific group and teacher
    def get_status_entry(group_id, teacher_id):
        return ProjectStatus.query.filter_by(
            group_id=group_id,
            teacher_id=teacher_id
        ).first()
    
    return render_template(
        'teacher_assigned_groups.html',
        groups=groups,
        pending_count=pending_count,
        accepted_count=accepted_count + conditionally_accepted_count,
        deferred_count=deferred_count,
        upcoming_vivas=upcoming_vivas,
        get_status_entry=get_status_entry
    )



@app.route('/teacher/viva_history')
@login_required
def teacher_viva_history():
    """View history of viva evaluations and feedback"""
    # Ensure the user is a teacher
    if current_user.role != 'teacher':
        flash('Access denied. This page is only for teachers.', 'danger')
        return redirect(url_for('index'))
    
    # Get status entries with filtering options
    query = ProjectStatus.query.filter_by(teacher_id=current_user.id)
    
    # Apply filters if provided
    status = request.args.get('status')
    if status:
        query = query.filter_by(status=status)
    
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    if date_from:
        date_from = datetime.datetime.strptime(date_from, '%Y-%m-%d')
        query = query.filter(ProjectStatus.updated_at >= date_from)
    
    if date_to:
        date_to = datetime.datetime.strptime(date_to, '%Y-%m-%d')
        date_to = date_to + datetime.timedelta(days=1)  # Include the end date
        query = query.filter(ProjectStatus.updated_at <= date_to)
    
    search = request.args.get('search')
    if search:
        # Join with StudentGroup to search by group ID or project title
        query = query.join(StudentGroup).filter(
            or_(
                StudentGroup.group_id.ilike(f'%{search}%'),
                StudentGroup.project_title.ilike(f'%{search}%')
            )
        )
    
    # Order by most recent first
    status_entries = query.order_by(ProjectStatus.updated_at.desc()).all()
    
    return render_template(
        'teacher_viva_history.html',
        status_entries=status_entries
    )

@app.route('/supervisor/add_project', methods=['POST'])
@login_required
def supervisor_add_project():
    # Only admins can create groups/projects
    if current_user.role != "admin":
        flash("Access denied. Only admins can create projects.", 'danger')
        return redirect(url_for('index'))
    
    group_id = request.form.get('group_id')
    project_title = request.form.get('project_title')
    project_description = request.form.get('project_description')
    
    # Check if group_id already exists
    if StudentGroup.query.filter_by(group_id=group_id).first():
        flash(f'Group ID {group_id} already exists', 'danger')
        return redirect(url_for('dashboard'))
    
    # Create new project group with current supervisor
    group = StudentGroup(
        group_id=group_id,
        project_title=project_title,
        project_description=project_description,
        supervisor_id=current_user.id
    )
    
    db.session.add(group)
    db.session.commit()
    
    flash(f'Project "{project_title}" added successfully', 'success')
    return redirect(url_for('dashboard'))

@app.route('/add_project_and_group', methods=['GET', 'POST'])
@login_required
def add_project_and_group():
    """Combined form for adding a project and assigning students to the group"""
    if current_user.role != "admin":
        flash("Access denied. Only admins can create projects and groups.", 'danger')
        return redirect(url_for('index'))
    
    # Get all students who are already assigned to groups
    assigned_students = db.session.query(GroupMember.user_id).all()
    assigned_ids = [student[0] for student in assigned_students]
    
    # Get only available students (not already in any group)
    available_students = User.query.filter(User.role == 'student')
    if assigned_ids:
        available_students = available_students.filter(~User.id.in_(assigned_ids))
    
    students = available_students.all()
    
    # Get all supervisors for dropdown
    supervisors = User.query.filter(User.role.in_(['supervisor', 'faculty', 'teacher'])).all()
    
    if request.method == 'POST':
        group_id = request.form.get('group_id')
        project_title = request.form.get('project_title')
        project_description = request.form.get('project_description')
        project_major = request.form.get('project_major')
        selected_students = request.form.getlist('selected_students')
        supervisor_id = request.form.get('supervisor_id')
        
        # Validate inputs
        if not group_id or not project_title:
            flash('Group ID and Project Title are required', 'danger')
            return render_template('add_project_and_group.html', students=students, supervisors=supervisors)
        
        # Validate supervisor is selected
        if not supervisor_id:
            flash('You must assign a supervisor to the project', 'danger')
            return render_template('add_project_and_group.html', students=students, supervisors=supervisors)
        
        # Check if group_id already exists
        if StudentGroup.query.filter_by(group_id=group_id).first():
            flash(f'Group ID {group_id} already exists', 'danger')
            return render_template('add_project_and_group.html', students=students, supervisors=supervisors)
        
        # Verify supervisor exists
        supervisor = User.query.get(int(supervisor_id))
        if not supervisor:
            flash('Selected supervisor does not exist', 'danger')
            return render_template('add_project_and_group.html', students=students, supervisors=supervisors)
        
        # Create new project group with selected supervisor
        group = StudentGroup(
            group_id=group_id,
            project_title=project_title,
            project_description=project_description,
            supervisor_id=int(supervisor_id)
        )
        
        db.session.add(group)
        db.session.commit()
        
        # Add selected students to the group
        for student_id in selected_students:
            group_member = GroupMember(
                user_id=int(student_id),
                group_id=group.id
            )
            db.session.add(group_member)
        
        db.session.commit()
        
        flash(f'Project "{project_title}" created and {len(selected_students)} students assigned successfully', 'success')
        
        if current_user.role == 'supervisor':
            return redirect(url_for('dashboard'))
        else:
            return redirect(url_for('dashboard'))
    
    return render_template('add_project_and_group.html', students=students, supervisors=supervisors)

@app.route('/supervisor/edit_project/<int:project_id>', methods=['POST'])
@login_required
def supervisor_edit_project(project_id):
    if current_user.role != "supervisor":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    group = StudentGroup.query.get_or_404(project_id)
    
    # Verify the supervisor owns this project
    if group.supervisor_id != current_user.id:
        flash("You don't have permission to edit this project.", 'danger')
        return redirect(url_for('dashboard'))
    
    group.group_id = request.form.get('group_id')
    group.project_title = request.form.get('project_title')
    group.project_description = request.form.get('project_description')
    
    db.session.commit()
    flash(f'Project "{group.project_title}" updated successfully', 'success')
    return redirect(url_for('dashboard'))

@app.route('/supervisor/delete_project/<int:project_id>', methods=['POST'])
@login_required
def supervisor_delete_project(project_id):
    if current_user.role != "supervisor":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    group = StudentGroup.query.get_or_404(project_id)
    
    # Verify the supervisor owns this project
    if group.supervisor_id != current_user.id:
        flash("You don't have permission to delete this project.", 'danger')
        return redirect(url_for('dashboard'))
    
    project_title = group.project_title
    
    try:
        # Delete associated members and remarks
        GroupMember.query.filter_by(group_id=group.id).delete()
        Remark.query.filter_by(group_id=group.id).delete()
        
        # Delete the group project
        db.session.delete(group)
        db.session.commit()
        
        flash(f'Group project "{project_title}" has been deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting group project: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/supervisor/assign_member', methods=['POST'])
@login_required
def supervisor_assign_member():
    if current_user.role != "supervisor":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    group_id = request.form.get('group_id')
    student_id = request.form.get('student_id')
    
    # Validate inputs
    if not group_id or not student_id:
        flash('Missing required information', 'danger')
        return redirect(url_for('dashboard'))
    
    # Find the group
    group = StudentGroup.query.filter_by(id=group_id).first()
    if not group:
        flash('Group not found', 'danger')
        return redirect(url_for('dashboard'))
    
    # Verify the supervisor owns this project
    if group.supervisor_id != current_user.id:
        flash("You don't have permission to modify this project.", 'danger')
        return redirect(url_for('dashboard'))
    
    # Find the student
    student = User.query.filter_by(id=student_id, role='student').first()
    if not student:
        flash('Student not found', 'danger')
        return redirect(url_for('dashboard'))
    
    # Check if the student is already in another group
    existing_membership = GroupMember.query.filter_by(user_id=student.id).first()
    if existing_membership and existing_membership.group_id != int(group_id):
        flash(f'Student is already assigned to another group', 'danger')
        return redirect(url_for('dashboard'))
    
    # Check if the group already has 2 members (maximum allowed)
    current_members = GroupMember.query.filter_by(group_id=group.id).count()
    if current_members >= 2:
        flash(f'Group already has the maximum of 2 members', 'danger')
        return redirect(url_for('dashboard'))
    
    # Create a GroupMember relationship if it doesn't already exist
    if not GroupMember.query.filter_by(user_id=student.id, group_id=group.id).first():
        group_member = GroupMember(user_id=student.id, group_id=group.id)
        db.session.add(group_member)
        db.session.commit()
        flash(f'Student {student.first_name} {student.last_name} added to group {group.group_id}', 'success')
    else:
        flash(f'Student already in the group', 'warning')
    
    return redirect(url_for('dashboard'))

@app.route('/supervisor/remove_member', methods=['POST'])
@login_required
def supervisor_remove_member():
    if current_user.role != "supervisor":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    group_id = request.form.get('group_id')
    student_id = request.form.get('student_id')
    
    # Validate inputs
    if not group_id or not student_id:
        flash('Missing required information', 'danger')
        return redirect(url_for('dashboard'))
    
    # Find the group
    group = StudentGroup.query.get_or_404(group_id)
    
    # Verify the supervisor owns this project
    if group.supervisor_id != current_user.id:
        flash("You don't have permission to modify this project.", 'danger')
        return redirect(url_for('dashboard'))
    
    # Delete the membership
    membership = GroupMember.query.filter_by(user_id=student_id, group_id=group_id).first()
    if membership:
        db.session.delete(membership)
        db.session.commit()
        flash('Student removed from group', 'success')
    else:
        flash('Student not found in group', 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/supervisor/group_members/<int:group_id>')
@login_required
def supervisor_group_members(group_id):
    if current_user.role != "supervisor":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    group = StudentGroup.query.get_or_404(group_id)
    
    # Verify the supervisor owns this project
    if group.supervisor_id != current_user.id:
        return jsonify([])
    
    # Get all members via relationship
    members = User.query.join(GroupMember).filter(GroupMember.group_id == group.id).all()
    
    return jsonify([{
        'id': member.id,
        'name': f"{member.first_name} {member.last_name}",
        'email': member.email
    } for member in members])

@app.route('/supervisor/available_students')
@login_required
def supervisor_available_students():
    if current_user.role != "supervisor":
        flash("Access denied.")
        return redirect(url_for('index'))
    
    # Get all students who are not assigned to any group
    assigned_students = db.session.query(GroupMember.user_id).all()
    assigned_ids = [student[0] for student in assigned_students]
    
    available_students = User.query.filter(User.role == 'student')
    if assigned_ids:
        available_students = available_students.filter(~User.id.in_(assigned_ids))
    
    students = available_students.all()
    
    return jsonify([{
        'id': student.id,
        'name': f"{student.first_name} {student.last_name}",
        'email': student.email
    } for student in students])

# Report Routes
@app.route('/admin/generate_user_summary', methods=['GET'])
@login_required
def generate_user_summary():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))
    
    # Generate user summary report in Excel format
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from flask import Response
    
    wb = Workbook()
    ws = wb.active
    ws.title = "User Summary"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary section
    all_users = User.query.all()
    students_count = len([u for u in all_users if u.role == 'student'])
    faculty_count = len([u for u in all_users if u.role == 'faculty'])
    supervisors_count = len([u for u in all_users if u.role == 'supervisor'])
    admins_count = len([u for u in all_users if u.role == 'admin'])
    
    # Add title
    ws['A1'] = "USER SUMMARY REPORT"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('A1:B1')
    ws.row_dimensions[1].height = 25
    
    # Summary statistics
    row = 3
    ws[f'A{row}'] = "User Role"
    ws[f'B{row}'] = "Count"
    
    for cell in [ws[f'A{row}'], ws[f'B{row}']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    summary_data = [
        ('Students', students_count),
        ('Faculty', faculty_count),
        ('Supervisors', supervisors_count),
        ('Admins', admins_count),
        ('TOTAL', len(all_users))
    ]
    
    for role, count in summary_data:
        ws[f'A{row}'] = role
        ws[f'B{row}'] = count
        
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            if role == 'TOTAL':
                cell.font = summary_font
                cell.fill = summary_fill
        
        row += 1
    
    # Add detailed user list
    row += 2
    ws[f'A{row}'] = "DETAILED USER LIST"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws.merge_cells(f'A{row}:H{row}')
    ws.row_dimensions[row].height = 20
    
    row += 1
    headers = ['ID', 'Email', 'First Name', 'Last Name', 'Role', 'Program', 'Semester', 'Created Date']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Add user data
    row += 1
    for user in all_users:
        ws.cell(row=row, column=1).value = user.id
        ws.cell(row=row, column=2).value = user.email
        ws.cell(row=row, column=3).value = user.first_name
        ws.cell(row=row, column=4).value = user.last_name
        ws.cell(row=row, column=5).value = user.role.upper() if user.role else ""
        ws.cell(row=row, column=6).value = user.program or ""
        ws.cell(row=row, column=7).value = user.semester or ""
        ws.cell(row=row, column=8).value = user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else ""
        
        # Apply borders and alternating colors
        for col_num in range(1, 9):
            cell = ws.cell(row=row, column=col_num)
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
    
    # Auto adjust column widths
    column_widths = [8, 25, 15, 15, 12, 12, 12, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=user_summary_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
    )

@app.route('/admin/generate_project_status', methods=['GET'])
@login_required
def generate_project_status():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))
    
    # Generate project status report in Excel format
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from flask import Response
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Status"
    
    # Define styles
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    summary_fill = PatternFill(start_color="E2EFD9", end_color="E2EFD9", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary section
    statuses = ['Pending', 'Accepted', 'Conditionally Accepted', 'Deferred']
    
    # Add title
    ws['A1'] = "PROJECT STATUS REPORT"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('A1:B1')
    ws.row_dimensions[1].height = 25
    
    # Status summary
    row = 3
    ws[f'A{row}'] = "Status"
    ws[f'B{row}'] = "Count"
    
    for cell in [ws[f'A{row}'], ws[f'B{row}']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    total_projects = 0
    
    # Get the latest status per group for accurate counting
    all_groups = StudentGroup.query.all()
    all_ps = ProjectStatus.query.order_by(ProjectStatus.created_at.desc()).all()
    latest_statuses = {}
    for ps in all_ps:
        if ps.group_id not in latest_statuses:
            latest_statuses[ps.group_id] = ps.status
    # Groups without any status record count as 'Pending'
    for g in all_groups:
        if g.id not in latest_statuses:
            latest_statuses[g.id] = 'Pending'
    
    for status in statuses:
        count = list(latest_statuses.values()).count(status)
        total_projects += count
        
        ws[f'A{row}'] = status
        ws[f'B{row}'] = count
        
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
    
    # Total
    ws[f'A{row}'] = "TOTAL PROJECTS"
    ws[f'B{row}'] = total_projects
    
    for cell in [ws[f'A{row}'], ws[f'B{row}']]:
        cell.border = thin_border
        cell.font = summary_font
        cell.fill = summary_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add detailed project list
    row += 2
    ws[f'A{row}'] = "PROJECT DETAILS"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{row}:F{row}')
    ws.row_dimensions[row].height = 20
    
    row += 1
    headers = ['Group ID', 'Project Title', 'Supervisor', 'Status', 'Feedback', 'Updated Date']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Add project data
    row += 1
    groups = StudentGroup.query.all()
    for group in groups:
        status_record = ProjectStatus.query.filter_by(group_id=group.id).first()
        supervisor_name = f"{group.supervisor.first_name} {group.supervisor.last_name}" if group.supervisor else "Unassigned"
        
        ws.cell(row=row, column=1).value = group.group_id
        ws.cell(row=row, column=2).value = group.project_title
        ws.cell(row=row, column=3).value = supervisor_name
        ws.cell(row=row, column=4).value = status_record.status if status_record else "No Status"
        ws.cell(row=row, column=5).value = status_record.feedback if status_record else ""
        ws.cell(row=row, column=6).value = status_record.updated_at.strftime('%Y-%m-%d %H:%M:%S') if status_record and status_record.updated_at else ""
        
        # Apply borders and alternating colors
        for col_num in range(1, 7):
            cell = ws.cell(row=row, column=col_num)
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="F0F8F0", end_color="F0F8F0", fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        row += 1
    
    # Auto adjust column widths
    column_widths = [12, 25, 20, 18, 30, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=project_status_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
    )

@app.route('/admin/generate_evaluation_summary', methods=['GET'])
@login_required
def generate_evaluation_summary():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))
    
    # Generate evaluation summary report in Excel format
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from flask import Response
    
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Evaluation Summary
    ws1 = wb.active
    ws1.title = "Evaluations"
    
    # Title
    ws1['A1'] = "EVALUATION SUMMARY REPORT"
    ws1['A1'].font = title_font
    ws1['A1'].fill = title_fill
    ws1.merge_cells('A1:F1')
    ws1.row_dimensions[1].height = 25
    
    # Headers for evaluations
    row = 3
    headers1 = ['Group ID', 'Project Title', 'Evaluator', 'Status', 'Feedback', 'Evaluation Date']
    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Add project evaluation data
    row = 4
    project_statuses = ProjectStatus.query.all()
    for ps in project_statuses:
        group = StudentGroup.query.get(ps.group_id)
        teacher = User.query.get(ps.teacher_id)
        
        if group and teacher:
            ws1.cell(row=row, column=1).value = group.group_id
            ws1.cell(row=row, column=2).value = group.project_title
            ws1.cell(row=row, column=3).value = f"{teacher.first_name} {teacher.last_name}"
            ws1.cell(row=row, column=4).value = ps.status
            ws1.cell(row=row, column=5).value = ps.feedback or ""
            ws1.cell(row=row, column=6).value = ps.created_at.strftime('%Y-%m-%d %H:%M:%S') if ps.created_at else ""
            
            # Apply borders and alternating colors
            for col_num in range(1, 7):
                cell = ws1.cell(row=row, column=col_num)
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            row += 1
    
    # Auto adjust columns for sheet 1
    column_widths1 = [12, 25, 20, 18, 30, 20]
    for col_num, width in enumerate(column_widths1, 1):
        ws1.column_dimensions[get_column_letter(col_num)].width = width
    
    # Sheet 2: Viva Examinations
    ws2 = wb.create_sheet("Viva Examinations")
    
    # Title
    ws2['A1'] = "VIVA EXAMINATION SCHEDULE"
    ws2['A1'].font = title_font
    ws2['A1'].fill = title_fill
    ws2.merge_cells('A1:G1')
    ws2.row_dimensions[1].height = 25
    
    # Headers for vivas
    row = 3
    headers2 = ['Group ID', 'Project Title', 'Examiner', 'Scheduled Date', 'Scheduled Time', 'Location', 'Status']
    for col_num, header in enumerate(headers2, 1):
        cell = ws2.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Add viva data
    row = 4
    vivas = Viva.query.all()
    for viva in vivas:
        group = StudentGroup.query.get(viva.group_id)
        examiner = User.query.get(viva.teacher_id)
        
        if group and examiner:
            ws2.cell(row=row, column=1).value = group.group_id
            ws2.cell(row=row, column=2).value = group.project_title
            ws2.cell(row=row, column=3).value = f"{examiner.first_name} {examiner.last_name}"
            ws2.cell(row=row, column=4).value = viva.scheduled_date.strftime('%Y-%m-%d') if viva.scheduled_date else ""
            ws2.cell(row=row, column=5).value = viva.scheduled_time
            ws2.cell(row=row, column=6).value = viva.location or ""
            ws2.cell(row=row, column=7).value = viva.status
            
            # Apply borders and alternating colors
            for col_num in range(1, 8):
                cell = ws2.cell(row=row, column=col_num)
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            row += 1
    
    # Auto adjust columns for sheet 2
    column_widths2 = [12, 25, 20, 15, 15, 20, 12]
    for col_num, width in enumerate(column_widths2, 1):
        ws2.column_dimensions[get_column_letter(col_num)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=evaluation_summary_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
    )

@app.route('/admin/export_user_data', methods=['GET'])
@login_required
def export_user_data():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))
    
    # Export user data in Excel format
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from flask import Response
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    headers = ['ID', 'Email', 'First Name', 'Last Name', 'Role', 'Program', 'Semester', 'Created Date']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Add data
    all_users = User.query.all()
    for row_num, user in enumerate(all_users, 2):
        ws.cell(row=row_num, column=1).value = user.id
        ws.cell(row=row_num, column=2).value = user.email
        ws.cell(row=row_num, column=3).value = user.first_name
        ws.cell(row=row_num, column=4).value = user.last_name
        ws.cell(row=row_num, column=5).value = user.role.upper() if user.role else ""
        ws.cell(row=row_num, column=6).value = user.program or ""
        ws.cell(row=row_num, column=7).value = user.semester or ""
        ws.cell(row=row_num, column=8).value = user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else ""
        
        # Apply borders and alternating colors
        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Auto adjust column widths
    column_widths = [8, 25, 15, 15, 12, 12, 12, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=user_data_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
    )

@app.route('/admin/export_project_data', methods=['GET'])
@login_required
def export_project_data():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))
    
    # Export project data in Excel format
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from flask import Response
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    
    # Define styles
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    headers = ['Group ID', 'Project Title', 'Description', 'Supervisor', 'Status', 'Feedback', 'Created', 'Updated']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Add data
    groups = StudentGroup.query.all()
    for row_num, group in enumerate(groups, 2):
        supervisor_name = f"{group.supervisor.first_name} {group.supervisor.last_name}" if group.supervisor else "Unassigned"
        status_record = ProjectStatus.query.filter_by(group_id=group.id).order_by(ProjectStatus.created_at.desc()).first()
        
        ws.cell(row=row_num, column=1).value = group.group_id
        ws.cell(row=row_num, column=2).value = group.project_title
        ws.cell(row=row_num, column=3).value = group.project_description or ""
        ws.cell(row=row_num, column=4).value = supervisor_name
        ws.cell(row=row_num, column=5).value = status_record.status if status_record else "No Status"
        ws.cell(row=row_num, column=6).value = status_record.feedback if status_record else ""
        ws.cell(row=row_num, column=7).value = group.created_at.strftime('%Y-%m-%d %H:%M:%S') if group.created_at else ""
        ws.cell(row=row_num, column=8).value = group.updated_at.strftime('%Y-%m-%d %H:%M:%S') if group.updated_at else ""
        
        # Apply borders and alternating colors
        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if row_num % 2 == 0:
                cell.fill = PatternFill(start_color="F0F8F0", end_color="F0F8F0", fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Auto adjust column widths
    column_widths = [12, 25, 30, 20, 18, 25, 20, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=project_data_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
    )

@app.route('/admin/export_evaluation_data', methods=['GET'])
@login_required
def export_evaluation_data():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))
    
    # Export evaluation data in Excel format
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    from flask import Response
    
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Project Evaluations
    ws1 = wb.active
    ws1.title = "Project Evaluations"
    
    headers1 = ['Group ID', 'Project Title', 'Evaluator', 'Status', 'Feedback', 'Evaluation Date']
    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Add project evaluation data
    project_statuses = ProjectStatus.query.all()
    for row_num, ps in enumerate(project_statuses, 2):
        group = StudentGroup.query.get(ps.group_id)
        teacher = User.query.get(ps.teacher_id)
        
        if group and teacher:
            ws1.cell(row=row_num, column=1).value = group.group_id
            ws1.cell(row=row_num, column=2).value = group.project_title
            ws1.cell(row=row_num, column=3).value = f"{teacher.first_name} {teacher.last_name}"
            ws1.cell(row=row_num, column=4).value = ps.status
            ws1.cell(row=row_num, column=5).value = ps.feedback or ""
            ws1.cell(row=row_num, column=6).value = ps.created_at.strftime('%Y-%m-%d %H:%M:%S') if ps.created_at else ""
            
            # Apply borders and alternating colors
            for col_num in range(1, 7):
                cell = ws1.cell(row=row_num, column=col_num)
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Auto adjust columns
    column_widths1 = [12, 25, 20, 18, 30, 20]
    for col_num, width in enumerate(column_widths1, 1):
        ws1.column_dimensions[get_column_letter(col_num)].width = width
    
    # Sheet 2: Viva Examinations
    ws2 = wb.create_sheet("Viva Examinations")
    
    headers2 = ['Group ID', 'Project Title', 'Examiner', 'Scheduled Date', 'Scheduled Time', 'Location', 'Status']
    for col_num, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Add viva data
    vivas = Viva.query.all()
    for row_num, viva in enumerate(vivas, 2):
        group = StudentGroup.query.get(viva.group_id)
        examiner = User.query.get(viva.teacher_id)
        
        if group and examiner:
            ws2.cell(row=row_num, column=1).value = group.group_id
            ws2.cell(row=row_num, column=2).value = group.project_title
            ws2.cell(row=row_num, column=3).value = f"{examiner.first_name} {examiner.last_name}"
            ws2.cell(row=row_num, column=4).value = viva.scheduled_date.strftime('%Y-%m-%d') if viva.scheduled_date else ""
            ws2.cell(row=row_num, column=5).value = viva.scheduled_time
            ws2.cell(row=row_num, column=6).value = viva.location or ""
            ws2.cell(row=row_num, column=7).value = viva.status
            
            # Apply borders and alternating colors
            for col_num in range(1, 8):
                cell = ws2.cell(row=row_num, column=col_num)
                cell.border = thin_border
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Auto adjust columns
    column_widths2 = [12, 25, 20, 15, 15, 20, 12]
    for col_num, width in enumerate(column_widths2, 1):
        ws2.column_dimensions[get_column_letter(col_num)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=evaluation_data_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
    )

@app.route('/admin/login_activity')
@login_required
def login_activity():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get login activity statistics
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Get all login attempts
    login_attempts = LoginAttempt.query.order_by(LoginAttempt.timestamp.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    # Get statistics
    total_logins = LoginAttempt.query.filter_by(success=True).count()
    total_failed = LoginAttempt.query.filter_by(success=False).count()
    unique_users = db.session.query(LoginAttempt.email).distinct().count()
    
    # Last 24 hours
    from datetime import datetime, timedelta
    last_24h = datetime.now() - timedelta(hours=24)
    logins_24h = LoginAttempt.query.filter(
        LoginAttempt.success == True,
        LoginAttempt.timestamp >= last_24h
    ).count()
    
    return render_template('login_activity.html',
                          login_attempts=login_attempts,
                          total_logins=total_logins,
                          total_failed=total_failed,
                          unique_users=unique_users,
                          logins_24h=logins_24h)

def recreate_tables():
    """Drop and recreate all tables."""
    db.drop_all()
    db.create_all()
    print("All tables recreated.")

# Render deployment initialization
if os.environ.get('RENDER'):
    with app.app_context():
        try:
            db.create_all()
            if not User.query.filter_by(role='admin').first():
                admin = User(email='admin@example.com', first_name='Admin', last_name='User', role='admin')
                admin.set_password('admin123')
                db.session.add(admin)
                db.session.commit()
                print("Render: Admin user created.")
        except Exception as e:
            print(f"Render Init Error: {e}")

# Vercel / Production deployment initialization (runs on cold start)
if os.environ.get('VERCEL') or os.environ.get('DATABASE_URL'):
    with app.app_context():
        try:
            db.create_all()
            # Seed default admin if missing
            if not User.query.filter_by(email='admin@example.com').first():
                admin = User(email='admin@example.com', first_name='Admin', last_name='User', role='admin')
                admin.set_password('admin123')
                db.session.add(admin)
                db.session.commit()
                print("Production: Admin user created.")
        except Exception as e:
            print(f"Production Init Error: {e}")

if __name__ == '__main__':
    with app.app_context():
        try:
            # Try to query the User model to check if all columns exist
            User.query.first()
            print("Database tables exist with current schema.")
            
            # Check if the Viva table has the correct schema
            try:
                Viva.query.order_by(Viva.scheduled_date, Viva.scheduled_time).first()
            except Exception as e:
                if "no such column: viva.scheduled_time" in str(e):
                    fix_viva_table()
            
            # Check if the group_member table exists - if not, recreate all tables
            inspector = inspect(db.engine)
            if 'group_member' not in inspector.get_table_names():
                print("Missing group_member table. Recreating all tables...")
                recreate_tables()
            else:
                # Only check if admin user exists if we didn't recreate tables
                admin = User.query.filter_by(email='admin@example.com').first()
                if not admin:
                    try:
                        # Create initial users
                        admin = User(email='admin@example.com', first_name='Admin', last_name='User', role='admin')
                        admin.set_password('admin123')
                        db.session.add(admin)
                        
                        # Create a sample teacher
                        teacher = User(email='teacher@example.com', first_name='John', last_name='Smith', role='teacher')
                        teacher.set_password('teacher123')
                        db.session.add(teacher)
                        
                        # Create a sample supervisor
                        supervisor = User(email='supervisor@example.com', first_name='David', last_name='Johnson', role='supervisor')
                        supervisor.set_password('supervisor123')
                        db.session.add(supervisor)
                        
                        # Create a sample student
                        student = User(email='student@example.com', first_name='Sarah', last_name='Johnson', role='student')
                        student.set_password('student123')
                        db.session.add(student)
                        
                        db.session.commit()
                        print("Sample users created.")
                    except Exception as e:
                        db.session.rollback()
                        print(f"Error creating sample users: {str(e)}")
            
        except Exception as e:
            # If there's an error (like missing columns), drop and recreate all tables
            print(f"Error initializing database: {str(e)}")
            recreate_tables()
    
    # Admin route for data integrity check
    @app.route('/admin/check-data-integrity')
    @login_required
    def check_data_integrity():
        if current_user.role != 'admin':
            flash('Access denied.', 'danger')
            return redirect(url_for('index'))
        
        try:
            issues = verify_data_integrity()
            
            if not issues:
                return jsonify({
                    'status': 'healthy',
                    'message': 'All data integrity checks passed!',
                    'total_issues': 0,
                    'issues': []
                })
            else:
                return jsonify({
                    'status': 'issues_found',
                    'message': f'Found {len(issues)} data integrity issues',
                    'total_issues': len(issues),
                    'issues': issues
                })
        except Exception as e:
            logger.error(f"Error checking data integrity: {str(e)}")
            return jsonify({
                'status': 'error',
                'message': f'Error checking integrity: {str(e)}',
                'total_issues': 0,
                'issues': []
            }), 500
    
# Register blueprints (must be after db/User definition to avoid circular imports)
# On Vercel, this must run at module level, not just in __main__
try:
    # Temporarily disabled due to circular import - will fix
    # from project_routes import project_bp
    # app.register_blueprint(project_bp, url_prefix='/projects')
    pass
except Exception as e:
    print(f"Error registering blueprint: {e}")

if __name__ == '__main__':
    with app.app_context():
        try:
            # Try to query the User model to check if all columns exist
            User.query.first()
            print("Database tables exist with current schema.")
            # ... (rest of the startup checks can remain here or be simplified)
            
            # Check if the group_member table exists - if not, recreate all tables
            inspector = inspect(db.engine)
            if 'group_member' not in inspector.get_table_names():
                print("Missing group_member table. Recreating all tables...")
                recreate_tables()

        except Exception as e:
            # If there's an error (like missing columns), drop and recreate all tables
            print(f"Error initializing database: {str(e)}")
            recreate_tables()

    app.run(host='0.0.0.0', debug=True)